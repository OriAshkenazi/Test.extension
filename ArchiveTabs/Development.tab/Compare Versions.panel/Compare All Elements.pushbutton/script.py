#! python3

import clr
import os
import sys
import time
import traceback
import io
import math
from collections import defaultdict
from typing import Tuple, Dict, List, Callable
from functools import wraps, lru_cache

clr.AddReference('RevitAPI')
clr.AddReference('RevitAPIUI')
# Revit 2019
# from Autodesk.Revit.DB import FilteredElementCollector, BuiltInCategory, BuiltInParameter, ElementId, Wall, Floor, Ceiling, FamilyInstance, Area, LocationCurve, RoofBase, UnitUtils, DisplayUnitType
# Revit 2023
from Autodesk.Revit import DB
from Autodesk.Revit.DB import Element, Document, FamilyInstance, FilteredElementCollector, BuiltInCategory, BuiltInParameter, ElementId, Wall, Floor, Ceiling, FamilyInstance, Area, LocationCurve, RoofBase, UnitUtils, ForgeTypeId
from Autodesk.Revit.UI import *
from System.Windows.Forms import FolderBrowserDialog, DialogResult
from System.Collections.Generic import List

import System
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment
from tqdm import tqdm

# Get the Revit application and document
uidoc = __revit__.ActiveUIDocument
doc = uidoc.Document

# List to keep track of all LRU cached functions
lru_cached_functions = []


class ElementMetricsCalculator:
    def __init__(self):
        self.metrics = {
            'length': 0.0,
            'area': 0.0,
            'volume': 0.0
        }
        self.errors = []
        self.logs = []
        self.category_handlers = {
            BuiltInCategory.OST_StructuralFoundation: self.handle_structural_foundation,
            BuiltInCategory.OST_Walls: self.handle_wall,
            BuiltInCategory.OST_Floors: self.handle_floor,
            BuiltInCategory.OST_Ceilings: self.handle_ceiling,
            BuiltInCategory.OST_Roofs: self.handle_roof,
            BuiltInCategory.OST_Rooms: self.handle_room,
            BuiltInCategory.OST_Stairs: self.handle_stairs,
            BuiltInCategory.OST_Railings: self.handle_railing
        }

    def calculate_metrics(self, element: Element, doc: Document) -> Tuple[Dict[str, float], List[str], List[str]]:
        self.metrics = {key: 0.0 for key in self.metrics}
        self.errors = [f"Element ID: {element.Id.IntegerValue}"]
        self.logs = []

        try:
            category = element.Category
            if category:
                self.logs.append(f"Category: {category.Name}")
                handler = self.category_handlers.get(category.Id.IntegerValue, self.handle_default)
            else:
                self.errors.append("Category: None")
                handler = self.handle_default

            handler(element, doc)

            self.convert_units()
            self.log_metrics()

        except Exception as e:
            self.errors.append(f"Error calculating metrics: {str(e)}")
            import traceback
            self.errors.append(traceback.format_exc())

        return self.metrics, self.errors, self.logs

    def handle_structural_foundation(self, element: Element, doc: Document):
        self.logs.append("Handling structural foundation.")
        self.metrics['length'] = self.get_parameter_value(element, 'Foundation Depth')
        diameter = self.get_parameter_value(element, 'Diameter')
        
        if diameter == 0:
            width = self.get_parameter_value(element, 'Width')
            depth = self.get_parameter_value(element, 'Depth')
        else:
            width = depth = diameter
        
        if diameter > 0:
            self.metrics['area'] = math.pi * (diameter/2)**2
        elif width > 0 and depth > 0:
            self.metrics['area'] = width * depth
        
        if self.metrics['length'] > 0 and self.metrics['area'] > 0:
            self.metrics['volume'] = self.metrics['area'] * self.metrics['length']
        else:
            self.metrics['volume'] = self.get_parameter_value(element, 'Volume')
        
        if self.metrics['length'] == 0:
            try:
                geo_elem = element.get_Geometry(Options())
                if geo_elem:
                    bbox = geo_elem.GetBoundingBox()
                    self.metrics['length'] = bbox.Max.Z - bbox.Min.Z
            except Exception as e:
                self.errors.append(f"Error calculating foundation length: {str(e)}")

    def handle_wall(self, element: Element, doc: Document):
        self.logs.append("Handling wall.")
        location = element.Location
        if isinstance(location, LocationCurve):
            self.metrics['length'] = location.Curve.Length
        self.metrics['area'] = self.get_parameter_value(element, BuiltInParameter.HOST_AREA_COMPUTED)
        self.metrics['volume'] = self.get_parameter_value(element, BuiltInParameter.HOST_VOLUME_COMPUTED)

    def handle_floor(self, element: Element, doc: Document):
        self.logs.append("Handling floor.")
        self.handle_default(element, doc)

    def handle_ceiling(self, element: Element, doc: Document):
        self.logs.append("Handling ceiling.")
        self.handle_default(element, doc)

    def handle_roof(self, element: Element, doc: Document):
        self.logs.append("Handling roof.")
        self.handle_default(element, doc)

    def handle_room(self, element: Element, doc: Document):
        self.logs.append("Handling room.")
        self.metrics['area'] = self.get_parameter_value(element, BuiltInParameter.ROOM_AREA)
        self.metrics['volume'] = self.get_parameter_value(element, BuiltInParameter.ROOM_VOLUME)

    def handle_stairs(self, element: Element, doc: Document):
        self.logs.append("Handling stairs.")
        self.metrics['length'] = self.get_parameter_value(element, BuiltInParameter.STAIRS_ACTUAL_RUN_LENGTH)

    def handle_railing(self, element: Element, doc: Document):
        self.logs.append("Handling railing.")
        location = element.Location
        if isinstance(location, LocationCurve):
            self.metrics['length'] = location.Curve.Length

    def handle_default(self, element: Element, doc: Document):
        self.logs.append("Handling default category.")
        self.metrics['length'] = self.get_parameter_value(element, BuiltInParameter.INSTANCE_LENGTH_PARAM)
        self.metrics['area'] = self.get_parameter_value(element, BuiltInParameter.HOST_AREA_COMPUTED)
        self.metrics['volume'] = self.get_parameter_value(element, BuiltInParameter.HOST_VOLUME_COMPUTED)
        
        if all(v == 0 for v in self.metrics.values()):
            bbox = element.get_BoundingBox(None)
            if bbox:
                self.metrics['length'] = max(bbox.Max.X - bbox.Min.X, bbox.Max.Y - bbox.Min.Y, bbox.Max.Z - bbox.Min.Z)

    def get_parameter_value(self, element: Element, param_id):
        param = None
        if isinstance(param_id, BuiltInParameter):
            param = element.get_Parameter(param_id)
        elif isinstance(param_id, str):
            param = element.LookupParameter(param_id)
        
        if param and param.HasValue:
            if param.StorageType == StorageType.Double:
                return param.AsDouble()
            elif param.StorageType == StorageType.Integer:
                return float(param.AsInteger())
            elif param.StorageType == StorageType.String:
                try:
                    return float(param.AsString())
                except ValueError:
                    return 0.0
        return 0.0

    def convert_units(self):
        try:
            self.metrics['length'] = UnitUtils.ConvertFromInternalUnits(self.metrics['length'], 
                                    ForgeTypeId.FromString("autodesk.spec.aec:meters-1.0.1"))
            self.metrics['area'] = UnitUtils.ConvertFromInternalUnits(self.metrics['area'], 
                                    ForgeTypeId.FromString("autodesk.spec.aec:squareMeters-1.0.1"))
            self.metrics['volume'] = UnitUtils.ConvertFromInternalUnits(self.metrics['volume'], 
                                    ForgeTypeId.FromString("autodesk.spec.aec:cubicMeters-1.0.1"))
        except Exception as e:
            self.errors.append(f"Unit conversion error: {str(e)}")

    def log_metrics(self):
        for metric, value in self.metrics.items():
            self.logs.append(f"{metric.capitalize()}: {value:.2f}")

def tracked_lru_cache(*args, **kwargs):
    def decorator(func):
        cached_func = lru_cache(*args, **kwargs)(func)
        lru_cached_functions.append(cached_func)
        return cached_func
    return decorator

def clear_all_lru_caches():
    for func in lru_cached_functions:
        func.cache_clear()

def get_category_family_type_names(element: Element, doc: Document) -> tuple:
    '''
    Get the family and type names of an element, along with Type ID and additional type info.
    
    Args:
        element (Element): The element to get the names from.
        doc (Document): The Revit document.
    
    Returns:
        tuple: A tuple containing:
            - category_name (str): The name of the element's category.
            - family_name (str): The name of the element's family.
            - type_name (str): The name of the element's type.
            - type_id (str): The ID of the element's type.
            - additional_info (str): Additional type information.
    
    Raises:
        ValueError: If input parameters are not of the expected types.
        AttributeError: If expected attributes or methods are missing.
    '''
    if not isinstance(element, Element) or not isinstance(doc, Document):
        raise ValueError("Invalid input types. Expected Element and Document.")

    default_values = {
        "category_name": "Unknown Category",
        "family_name": "Unknown Family",
        "type_name": "Unknown Type",
        "type_id": "Unknown Type ID",
        "additional_info": "No additional info"
    }

    try:
        category = element.Category
        default_values["category_name"] = category.Name if category else default_values["category_name"]

        element_type = doc.GetElement(element.GetTypeId())
        if element_type:
            type_name_param = element_type.get_Parameter(BuiltInParameter.SYMBOL_NAME_PARAM)
            default_values["type_name"] = type_name_param.AsString() if type_name_param and type_name_param.HasValue else default_values["type_name"]
            default_values["type_id"] = str(element_type.Id.IntegerValue)

            if isinstance(element, FamilyInstance):
                family = element.Symbol.Family
                default_values["family_name"] = family.Name if family else default_values["family_name"]
            elif hasattr(element_type, 'FamilyName'):
                default_values["family_name"] = element_type.FamilyName

            if default_values["family_name"] == "Unknown Family" and category:
                default_values["family_name"] = category.Name

            default_values["additional_info"] = get_additional_type_info(element_type)

    except AttributeError as ae:
        print(f"AttributeError: {ae}. This might be due to unexpected element structure.")
    except Exception as e:
        print(f"Unexpected error occurred: {e}")

    return tuple(default_values.values())

@tracked_lru_cache(maxsize=None)
def get_additional_type_info(element_type):
    '''
    Get additional information about the element type.

    Args:
        element_type (Element): The element type to get additional info for.
    Returns:
        str: A string containing additional information about the element type.
    '''
    if element_type is None:
        return "No additional info"

    info_params = [
        BuiltInParameter.ALL_MODEL_TYPE_MARK,
        BuiltInParameter.ALL_MODEL_DESCRIPTION,
        BuiltInParameter.ALL_MODEL_MANUFACTURER,
        BuiltInParameter.ALL_MODEL_MODEL,
        BuiltInParameter.ALL_MODEL_INSTANCE_COMMENTS,
        BuiltInParameter.ALL_MODEL_TYPE_COMMENTS
    ]

    additional_info = []
    for param in info_params:
        param_value = element_type.get_Parameter(param)
        if param_value and param_value.HasValue:
            additional_info.append(f"{param_value.Definition.Name}: {param_value.AsString()}")

    return " | ".join(additional_info) if additional_info else "No additional info"

def get_element_parameters(element):
    parameters = {}
    category = element.Category
    if category:
        category_name = category.Name
        category_id = category.Id.IntegerValue
    else:
        category_name = "Unknown"
        category_id = -1

    # Common parameters for most categories
    common_params = [
        ("Element ID", BuiltInParameter.ID_PARAM),
        ("Area", BuiltInParameter.HOST_AREA_COMPUTED),
        ("Assembly Description", BuiltInParameter.UNIFORMAT_DESCRIPTION),
        ("Category", BuiltInParameter.ELEM_CATEGORY_PARAM),
        ("Description", BuiltInParameter.ALL_MODEL_DESCRIPTION),
        ("Family", BuiltInParameter.ELEM_FAMILY_PARAM),
        ("Family and Type", BuiltInParameter.ELEM_FAMILY_AND_TYPE_PARAM),
        ("Family Name", BuiltInParameter.ELEM_FAMILY_PARAM),
        ("Type", BuiltInParameter.ELEM_TYPE_PARAM),
        ("Type Comments", BuiltInParameter.ALL_MODEL_TYPE_COMMENTS),
        ("Type Mark", BuiltInParameter.ALL_MODEL_TYPE_MARK),
        ("Type Name", BuiltInParameter.SYMBOL_NAME_PARAM)
    ]

    for param_name, built_in_param in common_params:
        param = element.get_Parameter(built_in_param)
        if param and param.HasValue:
            parameters[param_name] = param.AsString()

    # Category-specific parameters
    if category_name == "Doors" or category_name == "Windows":
        height_param = element.get_Parameter(BuiltInParameter.INSTANCE_HEIGHT_PARAM)
        if height_param and height_param.HasValue:
            parameters["Height"] = height_param.AsDouble()

    elif category_name == "Floors" or category_name == "Structural Foundations":
        depth_param = element.get_Parameter(BuiltInParameter.FLOOR_ATTR_THICKNESS_PARAM)
        if depth_param and depth_param.HasValue:
            parameters["Depth"] = depth_param.AsDouble()

    elif category_name == "Structural Columns":
        depth_param = element.get_Parameter(BuiltInParameter.FAMILY_BASE_SYMBOL_DEPTH)
        if depth_param and depth_param.HasValue:
            parameters["Depth"] = depth_param.AsDouble()

    elif category_name == "Walls":
        length_param = element.get_Parameter(BuiltInParameter.CURVE_ELEM_LENGTH)
        if length_param and length_param.HasValue:
            parameters["Length"] = length_param.AsDouble()

    elif category_name == "Stairs":
        actual_run_length = element.get_Parameter(BuiltInParameter.STAIRS_ACTUAL_RUN_LENGTH)
        if actual_run_length and actual_run_length.HasValue:
            parameters["Actual Run Length"] = actual_run_length.AsDouble()

    elif category_name == "Railings":
        length_param = element.get_Parameter(BuiltInParameter.CURVE_ELEM_LENGTH)
        if length_param and length_param.HasValue:
            parameters["Length"] = length_param.AsDouble()

    elif category_name == "Plumbing Fixtures":
        overall_height = element.get_Parameter(BuiltInParameter.FAMILY_HEIGHT_PARAM)
        overall_length = element.get_Parameter(BuiltInParameter.FAMILY_WIDTH_PARAM)
        if overall_height and overall_height.HasValue:
            parameters["Overall Height"] = overall_height.AsDouble()
        if overall_length and overall_length.HasValue:
            parameters["Overall Length"] = overall_length.AsDouble()

    return parameters

def get_parameter_value(element, param_id):
    '''
    Helper function to get a parameter value safely.
    Args:
        element: The Revit element
        param_id: Either a BuiltInParameter or a string representing the parameter name
    Returns:
        float: The parameter value as a float, or 0.0 if not found or not a number
    '''
    param = None
    if isinstance(param_id, BuiltInParameter):
        param = element.get_Parameter(param_id)
    elif isinstance(param_id, str):
        param = element.LookupParameter(param_id)
    
    if param and param.HasValue:
        if param.StorageType == StorageType.Double:
            return param.AsDouble()
        elif param.StorageType == StorageType.Integer:
            return float(param.AsInteger())
        elif param.StorageType == StorageType.String:
            try:
                return float(param.AsString())
            except ValueError:
                return 0.0
    return 0.0

def calculate_element_metrics(element, doc) -> Tuple[Dict[str, float], List[str], List[str]]:
    '''
    Calculate the metrics for an element using the ElementMetricsCalculator class.

    Args:
        element (Element): The element to calculate metrics for.
        doc (Document): The Revit document.
    Returns:
        Tuple[Dict[str, float], List[str], List[str]]: A tuple containing:
            - A dictionary of calculated metrics.
            - A list of errors encountered during calculation.
            - A list of logs for tracking the processing steps.
    '''
    calculator = ElementMetricsCalculator()
    return calculator.calculate_metrics(element, doc)

def get_type_metrics_old(doc):
    '''
    Get the metrics for all types in the document.

    Args:
        doc (Document): The Revit document.
    Returns:
        Tuple[Dict[Tuple[str, str], Dict[str, float]], List[str], Generator]: A tuple containing:
            - A dictionary of metrics for each type.
            - A list of global errors encountered during processing.
            - A generator to process elements and yield progress.
    '''
    metrics = defaultdict(lambda: {
        'count': 0, 'area': 0.0, 'volume': 0.0, 'length': 0.0,
        'type_id': '', 'additional_info': '', 'category': '', 
        'parameters': {}, 'errors': []
    })
    global_errors = []
    processed_elements = 0
    
    total_elements = FilteredElementCollector(doc).WhereElementIsNotElementType().GetElementCount()
    
    def element_processor():
        nonlocal processed_elements
        nonlocal global_errors
        for element in FilteredElementCollector(doc).WhereElementIsNotElementType().ToElements():
            element_errors = []
            try:
                category_name, family_name, type_name, type_id, additional_info = get_category_family_type_names(element, doc)
                key = (family_name, type_name)
                
                metrics[key]['count'] += 1
                metrics[key]['type_id'] = type_id
                metrics[key]['additional_info'] = additional_info
                metrics[key]['category'] = category_name
                
                element_metrics, calc_errors = calculate_element_metrics(element, doc)
                element_errors.extend(calc_errors)
                
                for metric, value in element_metrics.items():
                    if value is not None:
                        metrics[key][metric] += value
                
                # Add parameter information
                element_parameters = get_element_parameters(element)
                for param, value in element_parameters.items():
                    if param not in metrics[key]['parameters']:
                        metrics[key]['parameters'][param] = set()
                    metrics[key]['parameters'][param].add(value)
                
                if element_errors:
                    metrics[key]['errors'].extend(element_errors)
                
            except Exception as e:
                error_msg = f"Error processing element {element.Id} from document '{doc.Title}': {str(e)}\n{traceback.format_exc()}"
                element_errors.append(error_msg)
                global_errors.append(error_msg)
            
            finally:
                processed_elements += 1
                yield processed_elements, total_elements, element_errors
    
    return metrics, global_errors, element_processor()

def get_type_metrics(doc: Document):
    calculator = ElementMetricsCalculator()
    metrics = defaultdict(lambda: {
        'count': 0, 'area': 0.0, 'volume': 0.0, 'length': 0.0,
        'type_id': '', 'additional_info': '', 'category': '', 
        'parameters': {}, 'errors': []
    })
    global_errors = []
    
    total_elements = FilteredElementCollector(doc).WhereElementIsNotElementType().GetElementCount()
    
    def element_processor():
        processed_elements = 0
        for element in FilteredElementCollector(doc).WhereElementIsNotElementType().ToElements():
            try:
                category_name, family_name, type_name, type_id, additional_info = get_category_family_type_names(element, doc)
                key = (family_name, type_name)
                
                metrics[key]['count'] += 1
                metrics[key]['type_id'] = type_id
                metrics[key]['additional_info'] = additional_info
                metrics[key]['category'] = category_name
                
                element_metrics, calc_errors = calculator.calculate_metrics(element, doc)
                metrics[key]['errors'].extend(calc_errors)
                
                for metric, value in element_metrics.items():
                    if value is not None:
                        metrics[key][metric] += value
                
                element_parameters = get_element_parameters(element)
                for param, value in element_parameters.items():
                    if param not in metrics[key]['parameters']:
                        metrics[key]['parameters'][param] = set()
                    metrics[key]['parameters'][param].add(value)
                
            except Exception as e:
                error_msg = f"Error processing element {element.Id} from document '{doc.Title}': {str(e)}\n{traceback.format_exc()}"
                global_errors.append(error_msg)
            
            finally:
                processed_elements += 1
                yield processed_elements, total_elements
    
    return metrics, global_errors, element_processor()

def process_model(doc, model_name):
    '''
    Process the model to get the metrics and errors.

    Args:
        doc (Document): The Revit document to process.
        model_name (str): The name of the model.
    Returns:
        Tuple[Dict[Tuple[str, str], Dict[str, float]], List[str]]: A tuple containing:
            - A dictionary of metrics for each type.
            - A list of all errors encountered during processing.
    '''
    metrics, global_errors, processor = get_type_metrics(doc)
    total_elements = next(processor)[1]  # Get total elements from first yield
    
    all_errors = global_errors.copy()  # Create a copy of global_errors
    
    with tqdm(total=total_elements, desc=f"Processing {model_name} model") as pbar:
        for processed, total, element_errors in processor:
            all_errors.extend(element_errors)
            pbar.update(1)
    
    metrics = set_to_list(metrics)
    return metrics, all_errors

def compare_models(current_doc, old_doc_path):
    '''
    Compare two Revit models and generate a report.

    Args:
        current_doc (Document): The current Revit document.
        old_doc_path (str): The path to the old Revit document.
    Returns:
        Tuple[List[Dict[str, any]], List[str], set]: A tuple containing:
            - A list of dictionaries containing the comparison data.
            - A list of errors encountered during processing.
            - A set of all types processed.
    '''
    app = current_doc.Application
    all_errors = []

    print("Processing current model...")
    current_metrics, current_errors = process_model(current_doc, "current")
    all_errors.extend(current_errors)

    print("\nProcessing old model...")
    old_doc = None
    try:
        old_doc = app.OpenDocumentFile(old_doc_path)
        old_metrics, old_errors = process_model(old_doc, "old")
        all_errors.extend(old_errors)
    except Exception as e:
        all_errors.append(f"Error opening or processing old document: {str(e)}")
        return [], all_errors
    finally:
        if old_doc:
            old_doc.Close(False)

    if not current_metrics or not old_metrics:
        all_errors.append("Error: Failed to process one or both models.")
        return [], all_errors

    all_types = set(current_metrics.keys()) | set(old_metrics.keys())

    comparison_data = []
    for key in all_types:
        if key is None:
            continue
        family_name, type_name = key
        current = current_metrics.get(key, {'count': 0, 'length': 0, 'area': 0, 'volume': 0, 'category': 'N/A'})
        old = old_metrics.get(key, {'count': 0, 'length': 0, 'area': 0, 'volume': 0, 'category': 'N/A'})

        comparison_data.append({
            'Category': current.get('category', 'N/A'),
            'Family': family_name,
            'Type': type_name,
            'Type ID': current.get('type_id', 'N/A'),
            'Additional Info': current.get('additional_info', 'N/A'),
            'Old Count': old['count'],
            'Current Count': current['count'],
            'Count Diff': current['count'] - old['count'],
            'Count Diff %': (
                1 if old['count'] == 0 and current['count'] > 0 else
                -1 if old['count'] > 0 and current['count'] == 0 else
                ((current['count'] - old['count']) / old['count']) if old['count'] != 0 else 0
            ),
            'Old Length': old['length'],
            'Current Length': current['length'],
            'Length Diff': current['length'] - old['length'],
            'Length Diff %': (
                1 if old['length'] == 0 and current['length'] > 0 else
                -1 if old['length'] > 0 and current['length'] == 0 else
                ((current['length'] - old['length']) / old['length']) if old['length'] != 0 else 0
            ),
            'Old Area': old['area'],
            'Current Area': current['area'],
            'Area Diff': current['area'] - old['area'],
            'Area Diff %': (
                1 if old['area'] == 0 and current['area'] > 0 else
                -1 if old['area'] > 0 and current['area'] == 0 else
                ((current['area'] - old['area']) / old['area']) if old['area'] != 0 else 0
            ),
            'Old Volume': old['volume'],
            'Current Volume': current['volume'],
            'Volume Diff': current['volume'] - old['volume'],
            'Volume Diff %': (
                1 if old['volume'] == 0 and current['volume'] > 0 else
                -1 if old['volume'] > 0 and current['volume'] == 0 else
                ((current['volume'] - old['volume']) / old['volume']) if old['volume'] != 0 else 0
            )
        })

    return comparison_data, all_errors, all_types

def create_excel_report(comparison_data, output_path):
    '''
    Create an Excel report from the comparison data.

    Args:
        comparison_data (list): A list of dictionaries containing the comparison data.
        output_path (str): The path to save the Excel file.
    Returns:
        None
    '''
    # Sort the comparison_data before creating the Excel file
    sorted_data = sorted(comparison_data, key=lambda x: (x['Category'], x['Family'], x['Type']))
    
    wb = openpyxl.Workbook()
    ws_details = wb.active
    ws_details.title = "Type Comparison Details"
    
    headers = [
        "Category", "Family", "Type", "Type ID", "Additional Info",
        "Old Count", "Current Count", "Count Diff", "Count Diff %",
        "Old Len", "Current Len", "Len Diff", "Len Diff %",
        "Old Area", "Current Area", "Area Diff", "Area Diff %",
        "Old Vol", "Current Vol", "Vol Diff", "Vol Diff %",
        "Has Geometry"
    ]
    
    # Populate details worksheet
    for col, header in enumerate(headers, start=1):
        cell = ws_details.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    for row, data in enumerate(sorted_data, start=2):
        ws_details.cell(row=row, column=1, value=data['Category'])
        ws_details.cell(row=row, column=2, value=data['Family'])
        ws_details.cell(row=row, column=3, value=data['Type'])
        ws_details.cell(row=row, column=4, value=data['Type ID'])
        ws_details.cell(row=row, column=5, value=data['Additional Info'])
        
        # Count data
        ws_details.cell(row=row, column=6, value=data['Old Count'])
        ws_details.cell(row=row, column=7, value=data['Current Count'])
        ws_details.cell(row=row, column=8, value=data['Count Diff'])
        ws_details.cell(row=row, column=9, value=data['Count Diff %'])
        
        # Length data
        ws_details.cell(row=row, column=10, value=data['Old Length'])
        ws_details.cell(row=row, column=11, value=data['Current Length'])
        ws_details.cell(row=row, column=12, value=data['Length Diff'])
        ws_details.cell(row=row, column=13, value=data['Length Diff %'])
        
        # Area data
        ws_details.cell(row=row, column=14, value=data['Old Area'])
        ws_details.cell(row=row, column=15, value=data['Current Area'])
        ws_details.cell(row=row, column=16, value=data['Area Diff'])
        ws_details.cell(row=row, column=17, value=data['Area Diff %'])
        
        # Volume data
        ws_details.cell(row=row, column=18, value=data['Old Volume'])
        ws_details.cell(row=row, column=19, value=data['Current Volume'])
        ws_details.cell(row=row, column=20, value=data['Volume Diff'])
        ws_details.cell(row=row, column=21, value=data['Volume Diff %'])
        
        # Has Geometry
        has_geometry = any(data.get(metric) for metric in ('Old Length', 'Current Length', 'Old Area', 'Current Area', 'Old Volume', 'Current Volume'))
        ws_details.cell(row=row, column=22, value="Yes" if has_geometry else "No")

    # Apply number formats
    for row in ws_details.iter_rows(min_row=2, max_row=ws_details.max_row, min_col=6, max_col=21):
        for cell in row:
            if cell.column in [9, 13, 17, 21]:  # Percentage columns
                cell.number_format = '0%'
            else:
                cell.number_format = '0.0'

    # Adjust column widths for details worksheet
    for column in ws_details.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length
        ws_details.column_dimensions[column_letter].width = adjusted_width

    ws_details.auto_filter.ref = ws_details.dimensions

    wb.save(output_path)

def save_errors_to_file(errors, output_path):
    '''
    Save errors to a text file.

    Args:
        errors (list): A list of errors to save.
        output_path (str): The path to save the output file.
    '''
    try:
        with io.open(output_path, 'w', encoding='utf-8') as f:
            f.write("Errors occurred during processing:\n\n")
            if errors:
                for i, err in enumerate(errors, start=1):
                    try:
                        f.write(f"{i}. {err}\n")
                    except UnicodeEncodeError as e:
                        f.write(f"{i}. Error writing error message: {str(e)}\n")
            else:
                f.write("No errors were recorded during processing.\n")
        print(f"Errors have been saved to: {output_path}")
    except Exception as e:
        print(f"Failed to save errors to file: {str(e)}")
        print(f"Error details: {traceback.format_exc()}")

def save_all_types_to_file(all_types, output_path):
    '''
    Save all types to a text file.

    Args:
        all_types (set): A set of all types processed.
        output_path (str): The path to save the output file.
    '''
    with io.open(output_path, 'w', encoding='utf-8') as f:
        if all_types:
            f.write("All types processed:\n\n")
            for key in all_types:
                if key is None:
                    continue
                family_name, type_name = key
                f.write(f"{family_name} - {type_name}\n")

def get_file_path_flexible(prompt, file_extension):
    '''
    Get a file path from the user using various methods.

    Args:
        prompt (str): The prompt message.
        file_extension (str): The file extension to filter by.
    Returns:
        str: The file path selected by the user.
    '''
    # Try using Windows Forms
    try:
        from System.Windows.Forms import OpenFileDialog, DialogResult
        file_dialog = OpenFileDialog()
        file_dialog.Filter = f"{file_extension.upper()} Files (*.{file_extension})|*.{file_extension}"
        file_dialog.Title = prompt
        if file_dialog.ShowDialog() == DialogResult.OK:
            return file_dialog.FileName
    except:
        pass

    # Try using standard Python input
    try:
        return input(f"{prompt} (enter full path): ")
    except:
        pass

    # If all else fails, use TaskDialog for input
    try:
        path = TaskDialog.Show("File Path Input",
                               prompt,
                               TaskDialogCommonButtons.Ok | TaskDialogCommonButtons.Cancel,
                               TaskDialogResult.Ok)
        if path == TaskDialogResult.Ok:
            return TaskDialog.Show("File Path Input", "Enter the full file path:")
    except:
        pass

    # If we get here, all methods failed
    TaskDialog.Show("Error", "Unable to get file path input.")
    return None

def validate_file_path(file_path, should_exist=True):
    '''
    Validate the file path.

    Args:
        file_path (str): The file path to validate.
        should_exist (bool): Whether the file should exist.
    Returns:
        bool: True if the file path is valid, False otherwise.
    '''
    if not file_path:
        return False
    if should_exist and not os.path.exists(file_path):
        TaskDialog.Show("Error", f"The specified file does not exist: {file_path}")
        return False
    if not file_path.lower().endswith('.rvt') and not file_path.lower().endswith('.xlsx'):
        TaskDialog.Show("Error", f"Invalid file extension: {file_path}")
        return False
    return True

def get_folder_path(prompt):
    '''
    Get a folder path from the user using various methods.

    Args:
        prompt (str): The prompt message.
    Returns:
        str: The folder path selected by the user, or None if unsuccessful.
    '''
    # Try using Windows Forms
    try:
        from System.Windows.Forms import FolderBrowserDialog, DialogResult
        folder_dialog = FolderBrowserDialog()
        folder_dialog.Description = prompt
        if folder_dialog.ShowDialog() == DialogResult.OK:
            return folder_dialog.SelectedPath
    except:
        pass

    # Try using standard Python input
    try:
        folder_path = input(f"{prompt} (enter full path): ")
        if os.path.isdir(folder_path):
            return folder_path
        else:
            print("Invalid directory. Please try again.")
    except:
        pass

    # If all else fails, use TaskDialog for input
    try:
        path = TaskDialog.Show("Folder Path Input",
                               prompt,
                               TaskDialogCommonButtons.Ok | TaskDialogCommonButtons.Cancel,
                               TaskDialogResult.Ok)
        if path == TaskDialogResult.Ok:
            folder_path = TaskDialog.Show("Folder Path Input", "Enter the full folder path:")
            if os.path.isdir(folder_path):
                return folder_path
            else:
                TaskDialog.Show("Error", "Invalid directory. Please try again.")
    except:
        pass

    # If we get here, all methods failed
    TaskDialog.Show("Error", "Unable to get folder path input.")
    return None

def validate_folder_path(folder_path):
    '''
    Validate the folder path.

    Args:
        folder_path (str): The folder path to validate.
    Returns:
        bool: True if the folder path is valid, False otherwise.
    '''
    if not folder_path:
        return False
    if not os.path.exists(folder_path):
        TaskDialog.Show("Error", f"The specified folder does not exist: {folder_path}")
        return False
    return True

def set_to_list(metrics):
    for key in metrics:
        if 'parameters' in metrics[key]:
            metrics[key]['parameters'] = {k: list(v) for k, v in metrics[key]['parameters'].items()}
    return metrics

def main():
    clear_all_lru_caches()
    errors = []
    try:
        # Corrected file name handling
        file_name = doc.PathName.split("\\")[-1]  # Split by backslash and get the last element (file name with extension)
        file_name_base = file_name.split(".")[0]  # Remove the extension
        file_name_parts = file_name_base.split("_")  # Split by underscore
        if len(file_name_parts) >= 2:
            prefix = f"{file_name_parts[0]}_{file_name_parts[1]}"
        else:
            prefix = file_name_parts[0]  # Fallback to a simpler prefix

        # Get the old model path
        old_doc_path = get_file_path_flexible("Select the old Revit model file", "rvt")
        if not validate_file_path(old_doc_path):
            return

        # Get output folder path
        output_folder_path = get_folder_path("Select the output folder location")
        if not validate_folder_path(output_folder_path):
            return

        # Debugging: Print the selected folder path and prefix
        print(f"Selected output folder path: {output_folder_path}")
        print(f"Prefix: {prefix}")

        # Generate a timestamp for the file name
        timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        
        # Create file names with the timestamp
        output_file_name = f"Comparison.xlsx"
        errors_file_name = f"Comparison_errors.txt"
        all_types_file_name = f"All_Types.txt"
        
        # Create the full file paths
        full_output_dir = os.path.join(output_folder_path, prefix, timestamp)
        output_path = os.path.join(full_output_dir, output_file_name)
        errors_path = os.path.join(full_output_dir, errors_file_name)
        all_types_path = os.path.join(full_output_dir, all_types_file_name)

        # Create directories if they don't exist
        os.makedirs(full_output_dir, exist_ok=True)

        # Debugging: Print the full file paths
        print(f"Output file path: {output_path}")
        print(f"Errors file path: {errors_path}")
        print(f"All types file path: {all_types_path}")

        print("Starting model comparison...")
        comparison_data, comparison_errors, all_types = compare_models(doc, old_doc_path)
        errors.extend(comparison_errors)

        print("Saving all the compared types to a file...")
        if all_types:
            save_all_types_to_file(all_types, all_types_path)

        if not comparison_data:
            errors.append("No comparison data was generated.")
        else:
            print("Creating Excel report...")
            create_excel_report(comparison_data, output_path)
            print(f"Comparison complete. The results have been saved to {output_path}")
            TaskDialog.Show("Comparison Complete", f"The comparison results have been saved to {output_path}")

    except Exception as e:
        error_msg = f"Error in main function: {str(e)}\n{traceback.format_exc()}"
        print(error_msg)
        errors.append(error_msg)
        TaskDialog.Show("Error", f"An unexpected error occurred: {str(e)}\n\nPlease check the errors file for more details.")
    
    finally:
        if errors:
            save_errors_to_file(errors, errors_path)
            TaskDialog.Show("Errors Occurred", f"Errors occurred during processing. Please check the errors file: {errors_path}")
        else:
            print("No errors occurred during processing.")

    clear_all_lru_caches()

if __name__ == '__main__':
    main()
