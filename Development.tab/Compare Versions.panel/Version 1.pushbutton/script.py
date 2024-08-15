#! python3

import clr
import os
import sys
import time
import traceback
import io
from collections import defaultdict
from functools import wraps, lru_cache

clr.AddReference('RevitAPI')
clr.AddReference('RevitAPIUI')
from Autodesk.Revit.DB import FilteredElementCollector, BuiltInCategory, BuiltInParameter, ElementId, Wall, Floor, Ceiling, FamilyInstance, Area, LocationCurve, RoofBase, UnitUtils, DisplayUnitType
from Autodesk.Revit.UI import *
from System.Windows.Forms import FolderBrowserDialog, DialogResult
from System.Collections.Generic import List

import System
import openpyxl
from openpyxl.styles import Font, Alignment
import datetime

# Get the Revit application and document
uidoc = __revit__.ActiveUIDocument
doc = uidoc.Document

# List to keep track of all LRU cached functions
lru_cached_functions = []

def tracked_lru_cache(*args, **kwargs):
    def decorator(func):
        cached_func = lru_cache(*args, **kwargs)(func)
        lru_cached_functions.append(cached_func)
        return cached_func
    return decorator

def clear_all_lru_caches():
    for func in lru_cached_functions:
        func.cache_clear()

def progress_tracker(total_elements):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            start_time = time.time()
            estimated_times = []
            use_estimated_time = False
           
            def update_progress(i, force_print=False):
                nonlocal use_estimated_time
               
                current_time = time.time()
                total_elapsed_time = current_time - start_time
                progress = i / total_elements * 100
               
                if i > 1:  # Skip estimation for the first element
                    estimated_time_remaining = (total_elapsed_time / i) * (total_elements - i)
                    estimated_times.append(estimated_time_remaining)
                   
                    # Check negative gradient over last 800 elements
                    if len(estimated_times) >= 800:
                        if all(estimated_times[-j] < estimated_times[-j-1] + 0.5 for j in range(1, 800)) and not use_estimated_time:
                            use_estimated_time = True
                            print("Switching to estimated time remaining...")
                        estimated_times.pop(0)  # Remove oldest estimate to maintain 800 elements
                   
                    time_remaining = estimated_time_remaining if use_estimated_time else max(0, 360 - total_elapsed_time)
                   
                    if force_print or i % 400 == 0 or i == total_elements:
                        print(f"Progress: {progress:.2f}% ({i}/{total_elements} elements) | "
                              f"Time elapsed: {total_elapsed_time:.2f}s | "
                              f"Time remaining: {time_remaining:.2f}s")

            # Create a generator that yields elements
            def element_counter():
                elements = FilteredElementCollector(args[0]).WhereElementIsNotElementType().ToElements()
                for i, element in enumerate(elements, 1):
                    yield element
                    update_progress(i, force_print=(i % 400 == 0 or i == total_elements))

            # Pass the generator to the wrapped function
            return func(*args, element_counter=element_counter(), **kwargs)
        return wrapper
    return decorator

def get_family_and_type_names(element, doc):
    '''
    Get the family and type names of an element, along with Type ID and additional type info.
    Args:
        element (Element): The element to get the names from.
        doc (Document): The Revit document.
    Returns:
        tuple: A tuple containing the family name, type name, type ID, and additional type info.
    '''
    try:
        family_name = "Unknown Family"
        type_name = "Unknown Type"
        type_id = "Unknown Type ID"
        additional_info = "No additional info"

        # Get element type
        element_type = doc.GetElement(element.GetTypeId())
        
        if element_type is not None:
            # Get type name
            type_name_param = element_type.get_Parameter(BuiltInParameter.SYMBOL_NAME_PARAM)
            if type_name_param and type_name_param.HasValue:
                type_name = type_name_param.AsString()
            
            # Get type ID
            type_id = str(element_type.Id.IntegerValue)

            # Get family name
            if isinstance(element, FamilyInstance):
                family = element.Symbol.Family
                if family:
                    family_name = family.Name
            elif hasattr(element_type, 'FamilyName'):
                family_name = element_type.FamilyName
        
        # If family name is still unknown, use category name
        if family_name == "Unknown Family" and element.Category:
            family_name = element.Category.Name

        # Get additional type info
        additional_info = get_additional_type_info(element_type)

        return family_name, type_name, type_id, additional_info

    except Exception as e:
        return "Error Family", "Error Type", "Error Type ID", "Error getting info"

@tracked_lru_cache(maxsize=None)
def get_additional_type_info(element_type):
    '''
    Get additional information about the element type.

    Args:
        element_type (Element): The element type to get additional info for.
    Returns:
        str: A string containing additional information about the element type.
    '''
    if element_type is None:
        return "No additional info"

    info_params = [
        BuiltInParameter.ALL_MODEL_TYPE_MARK,
        BuiltInParameter.ALL_MODEL_DESCRIPTION,
        BuiltInParameter.ALL_MODEL_MANUFACTURER,
        BuiltInParameter.ALL_MODEL_MODEL
    ]

    additional_info = []
    for param in info_params:
        param_value = element_type.get_Parameter(param)
        if param_value and param_value.HasValue:
            additional_info.append(f"{param_value.Definition.Name}: {param_value.AsString()}")

    return " | ".join(additional_info) if additional_info else "No additional info"

def get_all_parameters(element):
    '''
    Get all parameters of an element as a list of strings.

    Args:
        element (Element): The element to get the parameters for.
    Returns:
        list: A list of strings containing the parameter names and values.
    '''
    params = []
    try:
        for param in element.Parameters:
            try:
                if param.HasValue:
                    value = param.AsValueString() or param.AsString() or str(param.AsDouble())
                else:
                    value = "No Value"
                params.append(f"{param.Definition.Name}: {value}")
            except Exception as e:
                params.append(f"{param.Definition.Name}: Error - {str(e)}")
    except Exception as e:
        params.append(f"Error getting parameters: {str(e)}")
    return params

def calculate_element_metrics(element, doc):
    '''
    Calculate the area, volume, and length of an element.

    Args:
        element (Element): The element to calculate the metrics for.
        doc (Document): The Revit document.
    Returns:
        dict: A dictionary containing the calculated metrics.
        list: A list of strings containing debug information.
    '''
    metrics = {'area': 0.0, 'length': 0.0, 'volume': 0.0}
    debug_info = [f"Element ID: {element.Id.IntegerValue}"]
   
    try:
        category = element.Category
        if category:
            debug_info.append(f"Category: {category.Name}")
            category_id = category.Id.IntegerValue
        else:
            debug_info.append("Category: None")
            category_id = -1

        # Try to get general parameters first
        area_param = element.get_Parameter(BuiltInParameter.HOST_AREA_COMPUTED)
        volume_param = element.get_Parameter(BuiltInParameter.HOST_VOLUME_COMPUTED)
        
        if area_param and area_param.HasValue:
            metrics['area'] = area_param.AsDouble()
        if volume_param and volume_param.HasValue:
            metrics['volume'] = volume_param.AsDouble()

        # Specific element handling
        if isinstance(element, Wall):
            location = element.Location
            if isinstance(location, LocationCurve):
                metrics['length'] = location.Curve.Length
        elif isinstance(element, (Floor, Ceiling, RoofBase)):
            pass  # Already handled by general parameters
        elif isinstance(element, FamilyInstance):
            bbox = element.get_BoundingBox(None)
            if bbox:
                metrics['length'] = max(bbox.Max.X - bbox.Min.X, bbox.Max.Y - bbox.Min.Y, bbox.Max.Z - bbox.Min.Z)
        elif category_id == int(BuiltInCategory.OST_Rooms):
            room_area_param = element.get_Parameter(BuiltInParameter.ROOM_AREA)
            room_volume_param = element.get_Parameter(BuiltInParameter.ROOM_VOLUME)
            if room_area_param and room_area_param.HasValue:
                metrics['area'] = room_area_param.AsDouble()
            if room_volume_param and room_volume_param.HasValue:
                metrics['volume'] = room_volume_param.AsDouble()
        elif category_id == int(BuiltInCategory.OST_Stairs):
            stair_length_param = element.get_Parameter(BuiltInParameter.STAIRS_ACTUAL_RUN_LENGTH)
            if stair_length_param and stair_length_param.HasValue:
                metrics['length'] = stair_length_param.AsDouble()
        elif category_id == int(BuiltInCategory.OST_Railings):
            location = element.Location
            if isinstance(location, LocationCurve):
                metrics['length'] = location.Curve.Length

        # Convert units (assuming input is in imperial units)
        metrics['area'] = UnitUtils.ConvertFromInternalUnits(metrics['area'], DisplayUnitType.DUT_SQUARE_METERS)
        metrics['volume'] = UnitUtils.ConvertFromInternalUnits(metrics['volume'], DisplayUnitType.DUT_CUBIC_METERS)
        metrics['length'] = UnitUtils.ConvertFromInternalUnits(metrics['length'], DisplayUnitType.DUT_METERS)

        debug_info.extend(get_all_parameters(element))

        for metric, value in metrics.items():
            debug_info.append(f"{metric.capitalize()}: {value:.2f}")

    except Exception as e:
        debug_info.append(f"Error calculating metrics: {str(e)}")
        debug_info.append(traceback.format_exc())

    return metrics, debug_info

@progress_tracker(total_elements=FilteredElementCollector(doc).WhereElementIsNotElementType().GetElementCount())
def get_type_metrics(doc, element_counter=None):
    '''
    Get the metrics for each unique element type in the document.

    Args:
        doc (Document): The Revit document.
        element_counter (generator): A generator that yields elements.
    Returns:
        dict: A dictionary containing the metrics for each unique element type.
        list: A list of strings containing errors encountered during processing.
        int: The total number of elements processed
    '''
    metrics = defaultdict(lambda: {'count': 0, 'area': 0.0, 'volume': 0.0, 'length': 0.0, 'type_id': '', 'additional_info': '', 'debug_info': []})
    
    errors = []
    processed_elements = 0
    
    for element in (element_counter or FilteredElementCollector(doc).WhereElementIsNotElementType().ToElements()[0:10000]):
        try:
            family_name, type_name, type_id, additional_info = get_family_and_type_names(element, doc)
            key = (family_name, type_name)
            
            metrics[key]['count'] += 1
            metrics[key]['type_id'] = type_id
            metrics[key]['additional_info'] = additional_info
            
            element_metrics, debug_info = calculate_element_metrics(element, doc)
            
            for metric, value in element_metrics.items():
                if value is not None:
                    metrics[key][metric] += value
            
            if not metrics[key]['debug_info']:
                metrics[key]['debug_info'] = debug_info
            
            processed_elements += 1

            # Limit to 5000 elements for testing
            # if processed_elements >= 5000:
            #     break
        
        except Exception as e:
            error_msg = f"Error processing element {element.Id}: {str(e)}\n{traceback.format_exc()}"
            if error_msg not in errors:
                errors.append(error_msg)
    
    return metrics, errors, processed_elements

def compare_models(current_doc, old_doc_path):
    '''
    Compare the types between the current and old models.

    Args:
        current_doc (Document): The current Revit document.
        old_doc_path (str): The path to the old Revit model file.
    Returns:
        tuple: A tuple containing a list of dictionaries with comparison data and a list of errors.
    '''
    app = current_doc.Application
    all_errors = []

    print("Processing current model...")
    current_metrics, current_errors, current_processed = get_type_metrics(current_doc)
    all_errors.extend(current_errors)
    print(f"Processed {current_processed} elements in the current model.")

    print("\nProcessing old model...")
    try:
        old_doc = app.OpenDocumentFile(old_doc_path)
        old_metrics, old_errors, old_processed = get_type_metrics(old_doc)
        all_errors.extend(old_errors)
        print(f"Processed {old_processed} elements in the old model.")
        old_doc.Close(False)
    except Exception as e:
        all_errors.append(f"Error opening old document: {str(e)}")
        return [], all_errors

    if current_metrics is None or old_metrics is None:
        all_errors.append("Error: Failed to process one or both models.")
        return [], all_errors

    all_types = set(current_metrics.keys()) | set(old_metrics.keys())

    comparison_data = []
    for key in all_types:
        if key is None:
            continue  # Skip None keys
        family_name, type_name = key
        current = current_metrics.get(key, {'count': 0, 'area': 0, 'volume': 0, 'length': 0})
        old = old_metrics.get(key, {'count': 0, 'area': 0, 'volume': 0, 'length': 0})

        comparison_data.append({
            'Family': family_name,
            'Type': type_name,
            'Type ID': current.get('type_id', 'N/A'),
            'Additional Info': current.get('additional_info', 'N/A'),
            'Current Count': current['count'],
            'Old Count': old['count'],
            'Count Diff': current['count'] - old['count'],
            'Current Area': current['area'],
            'Old Area': old['area'],
            'Area Diff': current['area'] - old['area'],
            'Current Volume': current['volume'],
            'Old Volume': old['volume'],
            'Volume Diff': current['volume'] - old['volume'],
            'Current Length': current['length'],
            'Old Length': old['length'],
            'Length Diff': current['length'] - old['length']
        })

    return comparison_data, all_errors, all_types

def create_excel_report(comparison_data, output_path):
    '''
    Create an Excel report from the comparison data.

    Args:
        comparison_data (list): A list of dictionaries containing the comparison data.
        output_path (str): The path to save the Excel file.
    Returns:
        None
    '''
    wb = openpyxl.Workbook()
    ws_details = wb.active
    ws_details.title = "Type Comparison Details"
    
    headers = [
        "Family", "Type", "Type ID", "Additional Info",
        "Current Count", "Old Count", "Count Diff",
        "Current Area (sqm)", "Old Area (sqm)", "Area Diff (sqm)",
        "Current Vol (cu m)", "Old Vol (cu m)", "Vol Diff (cu m)",
        "Current Len (m)", "Old Len (m)", "Len Diff (m)"
    ]
    
    # sorted_comparison_data = sorted(comparison_data, key=lambda x: safe_sort_key((x.get('Family'), x.get('Type'))))

    # Populate details worksheet
    for col, header in enumerate(headers, start=1):
        cell = ws_details.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # 'Family': family_name,
    # 'Type': type_name,
    # 'Type ID': current.get('type_id', 'N/A'),
    # 'Additional Info': current.get('additional_info', 'N/A'),
    # 'Current Count': current['count'],
    # 'Old Count': old['count'],
    # 'Count Diff': current['count'] - old['count'],
    # 'Current Area': current['area'],
    # 'Old Area': old['area'],
    # 'Area Diff': current['area'] - old['area'],
    # 'Current Volume': current['volume'],
    # 'Old Volume': old['volume'],
    # 'Volume Diff': current['volume'] - old['volume'],
    # 'Current Length': current['length'],
    # 'Old Length': old['length'],
    # 'Length Diff': current['length'] - old['length']

    for row, data in enumerate(comparison_data, start=2):
        ws_details.cell(row=row, column=1, value=data['Family'])
        ws_details.cell(row=row, column=2, value=data['Type'])
        ws_details.cell(row=row, column=3, value=data['Type ID'])
        ws_details.cell(row=row, column=4, value=data['Additional Info'])
        ws_details.cell(row=row, column=5, value=data['Current Count'])
        ws_details.cell(row=row, column=6, value=data['Old Count'])
        ws_details.cell(row=row, column=7, value=data['Count Diff'])
        ws_details.cell(row=row, column=8, value=data['Current Area'])
        ws_details.cell(row=row, column=9, value=data['Old Area'])
        ws_details.cell(row=row, column=10, value=data['Area Diff'])
        ws_details.cell(row=row, column=11, value=data['Current Volume'])
        ws_details.cell(row=row, column=12, value=data['Old Volume'])
        ws_details.cell(row=row, column=13, value=data['Volume Diff'])
        ws_details.cell(row=row, column=14, value=data['Current Length'])
        ws_details.cell(row=row, column=15, value=data['Old Length'])
        ws_details.cell(row=row, column=16, value=data['Length Diff'])
        
    
    # Adjust column widths for details worksheet
    for column in ws_details.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length
        ws_details.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)

def save_errors_to_file(errors, output_path):
    '''
    Save errors to a text file.

    Args:
        errors (list): A list of errors to save.
        output_path (str): The path to save the output file.
    '''
    with io.open(output_path, 'w', encoding='utf-8') as f:
        f.write("Errors occurred during processing:\n\n")
        if errors:
            for err, i in enumerate(errors, start=1):
                try:
                    f.write(f"{i}. {err}\n")
                except UnicodeEncodeError as e:
                    f.write(f"Error writing data for {key}: {str(e)}\n")

def save_all_types_to_file(all_types, output_path):
    '''
    Save all types to a text file.

    Args:
        all_types (set): A set of all types processed.
        output_path (str): The path to save the output file.
    '''
    with io.open(output_path, 'w', encoding='utf-8') as f:
        if all_types:
            f.write("All types processed:\n\n")
            for key in all_types:
                if key is None:
                    continue
                family_name, type_name = key
                f.write(f"{family_name} - {type_name}\n")

def get_file_path_flexible(prompt, file_extension):
    '''
    Get a file path from the user using various methods.

    Args:
        prompt (str): The prompt message.
        file_extension (str): The file extension to filter by.
    Returns:
        str: The file path selected by the user.
    '''
    # Try using Windows Forms
    try:
        from System.Windows.Forms import OpenFileDialog, DialogResult
        file_dialog = OpenFileDialog()
        file_dialog.Filter = f"{file_extension.upper()} Files (*.{file_extension})|*.{file_extension}"
        file_dialog.Title = prompt
        if file_dialog.ShowDialog() == DialogResult.OK:
            return file_dialog.FileName
    except:
        pass

    # Try using standard Python input
    try:
        return input(f"{prompt} (enter full path): ")
    except:
        pass

    # If all else fails, use TaskDialog for input
    try:
        path = TaskDialog.Show("File Path Input",
                               prompt,
                               TaskDialogCommonButtons.Ok | TaskDialogCommonButtons.Cancel,
                               TaskDialogResult.Ok)
        if path == TaskDialogResult.Ok:
            return TaskDialog.Show("File Path Input", "Enter the full file path:")
    except:
        pass

    # If we get here, all methods failed
    TaskDialog.Show("Error", "Unable to get file path input.")
    return None

def validate_file_path(file_path, should_exist=True):
    '''
    Validate the file path.

    Args:
        file_path (str): The file path to validate.
        should_exist (bool): Whether the file should exist.
    Returns:
        bool: True if the file path is valid, False otherwise.
    '''
    if not file_path:
        return False
    if should_exist and not os.path.exists(file_path):
        TaskDialog.Show("Error", f"The specified file does not exist: {file_path}")
        return False
    if not file_path.lower().endswith('.rvt') and not file_path.lower().endswith('.xlsx'):
        TaskDialog.Show("Error", f"Invalid file extension: {file_path}")
        return False
    return True

def get_folder_path(prompt):
    '''
    Get a folder path from the user using various methods.

    Args:
        prompt (str): The prompt message.
    Returns:
        str: The folder path selected by the user.
    '''
    try:
        from System.Windows.Forms import FolderBrowserDialog, DialogResult
        folder_dialog = FolderBrowserDialog()
        folder_dialog.Description = prompt
        if folder_dialog.ShowDialog() == DialogResult.OK:
            return folder_dialog.SelectedPath
    except:
        return TaskDialog.Show("Folder Path Input", prompt)

def validate_folder_path(folder_path):
    '''
    Validate the folder path.

    Args:
        folder_path (str): The folder path to validate.
    Returns:
        bool: True if the folder path is valid, False otherwise.
    '''
    if not folder_path:
        return False
    if not os.path.exists(folder_path):
        TaskDialog.Show("Error", f"The specified folder does not exist: {folder_path}")
        return False
    return True

def safe_sort_key(item):
    if item is None:
        return ("", "")  # Return empty tuple for None values
    if isinstance(item, tuple):
        return tuple(safe_sort_key(i) for i in item)
    return (str(item), "")

def main():
    clear_all_lru_caches()
    try:
        file_name = doc.PathName.split("/")[-1].split(".")[0].split("_")
        prefix = f"{file_name[0]}_{file_name[1]}"

        # Get the old model path
        old_doc_path = get_file_path_flexible("Select the old Revit model file", "rvt")
        if not validate_file_path(old_doc_path):
            return

        # Get output folder path
        output_folder_path = get_folder_path("Select the output folder location")
        if not validate_folder_path(output_folder_path):
            return

        # Generate a timestamp for the file name
        timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        
        # Create a new file name with the timestamp
        output_file_name = f"{prefix}_Comparison_{timestamp}.xlsx"

        # Create a new errors file name with the timestamp
        errors_file_name = f"{prefix}_Comparison_errors_{timestamp}.txt"
        
        # Create the full file path by joining the folder path and the new file name
        output_path = os.path.join(output_folder_path, output_file_name)

        # Create the full errors file path
        errors_path = os.path.join(output_folder_path, errors_file_name)

        all_types_path = os.path.join(output_folder_path, f"{prefix}_All_Types_{timestamp}.txt")

        print("Starting model comparison...")
        comparison_data, errors, all_types = compare_models(doc, old_doc_path)

        print("Saving all the compared types to a file...")
        if all_types: # Save all types to a file
            save_all_types_to_file(all_types, all_types_path)

        if errors: # Check for errors during processing
            save_errors_to_file(errors, all_types, errors_path)
            TaskDialog.Show("Comparison Failed", f"Errors occurred during processing. Please check the errors file.")

        if not comparison_data:
            TaskDialog.Show("Comparison Failed", "No comparison data was generated. Please check the errors.")
            return

        print("Creating Excel report...")
        create_excel_report(comparison_data, output_path)
        TaskDialog.Show("Comparison Complete", f"The comparison results have been saved to {output_path}")
        clear_all_lru_caches()

    except Exception as e:
        print(f"Error in main function: {str(e)}")
        print(traceback.format_exc())
        TaskDialog.Show("Error", f"An unexpected error occurred: {str(e)}\n\nPlease check the output window for more details.")
        clear_all_lru_caches()

# Call the main function
if __name__ == "__main__":
    main()
