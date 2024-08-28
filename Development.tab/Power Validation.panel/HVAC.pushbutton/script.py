#! python3

import clr
import os
import sys
import time
import traceback
import io
import math
from collections import defaultdict
from typing import Tuple, Dict, List, Callable
from functools import wraps, lru_cache

clr.AddReference('RevitAPI')
clr.AddReference('RevitAPIUI')
from Autodesk.Revit.DB import *
from Autodesk.Revit.UI import *
from System.Windows.Forms import FolderBrowserDialog, DialogResult
from System.Collections.Generic import List

import System
from System import Guid
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment
from tqdm import tqdm

# Get the Revit application and document
uidoc = __revit__.ActiveUIDocument
doc = uidoc.Document

lru_cached_functions = []

def tracked_lru_cache(*args, **kwargs):
    """
    Decorator to track functions using LRU cache.
    
    Args:
        *args: Positional arguments for lru_cache.
        **kwargs: Keyword arguments for lru_cache.
    
    Returns:
        Decorator function to wrap the target function.
    """
    def decorator(func):
        cached_func = lru_cache(*args, **kwargs)(func)
        lru_cached_functions.append(cached_func)
        return cached_func
    return decorator

def clear_all_lru_caches():
    """
    Clear all LRU caches for tracked functions.
    """
    for func in lru_cached_functions:
        func.cache_clear()

def get_linked_documents():
    """
    Retrieve linked documents in the current project.
    
    Returns:
        List[Document]: List of linked Revit documents.
    """
    linked_docs = []
    link_instances = FilteredElementCollector(doc).OfClass(RevitLinkInstance)

    for link_instance in link_instances:
        linked_doc = link_instance.GetLinkDocument()
        if linked_doc is not None:
            linked_docs.append(linked_doc)
    
    print(f"Found {len(linked_docs)} linked documents.")
    return linked_docs

def create_parameter_filters(doc, name, category_id, search_strings):
    category = doc.Settings.Categories.get_Item(category_id)
    if not category:
        raise ValueError(f"Category not found for id: {category_id}")

    print(f"Creating filters for category: {category.Name}")

    # Use "Family and Type" parameter instead of ELEM_FAMILY_PARAM
    family_and_type_param_id = ElementId(BuiltInParameter.ELEM_FAMILY_AND_TYPE_PARAM)
    
    filters = []
    for i, search_string in enumerate(search_strings.split(',')):
        if search_string.strip():
            try:
                rule = ParameterFilterRuleFactory.CreateContainsRule(family_and_type_param_id, search_string.strip(), False)
                element_filter = ElementParameterFilter(rule)
                filter_name = f"{name}_{i+1}"
                filter_element = ParameterFilterElement.Create(doc, filter_name, List[ElementId]([category_id]), element_filter)
                filters.append(filter_element)
                print(f"Created filter for search string: '{search_string.strip()}'")
            except Exception as e:
                print(f"Error creating filter for '{search_string.strip()}': {str(e)}")
    
    if not filters:
        raise ValueError(f"No valid filters created for category: {category.Name}")

    print(f"Created {len(filters)} filters")
    return filters
    
def get_3d_view_family_type(doc):
    view_family_types = FilteredElementCollector(doc).OfClass(ViewFamilyType).ToElements()
    for vft in view_family_types:
        if vft.ViewFamily == ViewFamily.ThreeDimensional:
            return vft.Id
    return None

def get_elements_from_linked_docs(doc, category):
    elements = []
    links = FilteredElementCollector(doc).OfClass(RevitLinkInstance).ToElements()
    for link in links:
        link_doc = link.GetLinkDocument()
        if link_doc:
            elements.extend(FilteredElementCollector(link_doc).OfCategory(category).WhereElementIsNotElementType().ToElements())
    return elements

def create_or_modify_3d_view(doc, power_devices, outlets, view_name="VDC - HVAC ELEC verification", is_new_view=False):
    print(f"{'Creating' if is_new_view else 'Modifying'} 3D view: {view_name}")
    t = Transaction(doc, f"{'Create' if is_new_view else 'Modify'} 3D View")
    try:
        t.Start()
        
        if is_new_view:
            view_family_type_id = get_3d_view_family_type(doc)
            if view_family_type_id is None or doc.GetElement(view_family_type_id).ViewFamily != ViewFamily.ThreeDimensional:
                raise ValueError("Invalid or missing 3D view family type.")
            
            new_3d_view = View3D.CreateIsometric(doc, view_family_type_id)
            try:
                new_3d_view.Name = view_name
            except Autodesk.Revit.Exceptions.ArgumentException:
                new_3d_view.Name = f"{view_name}_{Guid.NewGuid().ToString()}"
            view = new_3d_view
        else:
            view = next((v for v in FilteredElementCollector(doc).OfClass(View3D) 
                         if not v.IsTemplate and v.Name == view_name), None)
            if not view:
                raise ValueError(f"Existing view '{view_name}' not found.")

        # Apply view template
        view_template = create_or_get_view_template(doc, "VDC - HVAC ELEC verification")
        if view_template:
            view.ViewTemplateId = view_template.Id
        
        t.Commit()
        print(f"Successfully {'created' if is_new_view else 'modified'} 3D view: {view.Name}")
        return view
    except Exception as e:
        if t.HasStarted() and not t.HasEnded():
            t.RollBack()
        print(f"Error {'creating' if is_new_view else 'modifying'} 3D view: {str(e)}")
        print(traceback.format_exc())
        return None

def create_or_get_view_template(doc, template_name):
    existing_template = next((v for v in FilteredElementCollector(doc).OfClass(View3D)
                              if v.IsTemplate and v.Name == template_name), None)
    if existing_template:
        return existing_template

    t = Transaction(doc, "Create View Template")
    t.Start()
    try:
        view_family_type_id = next(vft.Id for vft in FilteredElementCollector(doc).OfClass(ViewFamilyType)
                                   if vft.ViewFamily == ViewFamily.ThreeDimensional)
        
        new_template = View3D.CreateIsometric(doc, view_family_type_id)
        new_template.Name = template_name
        new_template.ViewTemplateId = ElementId.InvalidElementId
        
        # Set view properties
        new_template.DetailLevel = ViewDetailLevel.Fine
        new_template.DisplayStyle = DisplayStyle.Realistic

        # Set visibility for categories
        visible_categories = [
            "Analysis Display Style", "Analysis Results", "Bridge Abutments", "Bridge Arches",
            "Bridge Bearings", "Bridge Cables", "Bridge Decks", "Bridge Foundations",
            "Bridge Girders", "Bridge Piers", "Bridge Towers", "Cable Tray Runs",
            "Conduit Runs", "Coordination Model", "Curtain Grids", "Duct Systems",
            "Electrical Circuits", "Electrical Fixtures", "Electrical Spare/Space Circuits",
            "Filled region", "HVAC Zones", "Imports in Families", "Masking Region",
            "Materials", "Mechanical Equipment", "Mechanical Equipment Sets",
            "Pipe Segments", "Piping Systems", "Project Information", "RVT Links",
            "Rebar Shape", "Rooms", "Routing Preferences", "Sheets", "Spaces",
            "Switch System", "Walls"
        ]

        all_categories = doc.Settings.Categories
        for category in all_categories:
            if category.CategoryType == CategoryType.Model:
                should_be_visible = category.Name in visible_categories
                try:
                    new_template.SetCategoryHidden(category.Id, not should_be_visible)
                except:
                    print(f"Could not set visibility for category: {category.Name}")

        # Set transparency for floors and walls
        override_settings = OverrideGraphicSettings()
        override_settings.SetSurfaceTransparency(90)
        new_template.SetCategoryOverrides(ElementId(BuiltInCategory.OST_Floors), override_settings)
        new_template.SetCategoryOverrides(ElementId(BuiltInCategory.OST_Walls), override_settings)

        # Set other properties
        new_template.Scale = 100
        new_template.Discipline = 4095  # This might need adjustment based on the Revit API enumeration
        new_template.get_Parameter(BuiltInParameter.VIEW_PARTS_VISIBILITY).Set(1)
        new_template.get_Parameter(BuiltInParameter.VIEW_SHOW_HIDDEN_LINES).Set(1)
        new_template.get_Parameter(BuiltInParameter.VIEW_SHOW_SUNPATH).Set(0)

        t.Commit()
        return new_template
    except Exception as e:
        t.RollBack()
        print(f"Error creating view template: {str(e)}")
        return None

def get_elements_by_type_ids_from_linked_docs(type_ids, linked_docs):
    """
    Fetch elements in the linked documents that match the specified type IDs.
    
    Args:
        type_ids (List[ElementId]): List of element type IDs to collect.
        linked_docs (List[Document]): List of linked Revit documents.
    
    Returns:
        List[Element]: List of elements that match the provided type IDs.
    """
    elements = []
    for linked_doc in linked_docs:
        collector = FilteredElementCollector(linked_doc).OfClass(FamilyInstance)
        for element in collector:
            if element.GetTypeId().IntegerValue in type_ids:
                elements.append(element)
    print(f"Found {len(elements)} elements in linked documents.")
    return elements

@tracked_lru_cache(maxsize=None)
def calculate_distance(source_point, target_point):
    """
    Calculate the distance between two points and convert it to meters.
    
    Args:
        source_point (XYZ): The source point.
        target_point (XYZ): The target point.
    
    Returns:
        float: The distance between the source and target points in meters.
    """
    distance_feet = source_point.DistanceTo(target_point)
    return distance_feet * 0.3048  # Convert feet to meters

def calculate_nearest_distance(source_elements, target_elements):
    results = []
    total_comparisons = len(source_elements) * len(target_elements)
    
    with tqdm(total=total_comparisons, desc="Calculating distances") as pbar:
        for source, source_doc, local_source in source_elements:
            source_point = source.Location.Point
            min_distance = float('inf')
            nearest_target = None
            nearest_target_doc = None
            nearest_local_target = None
            
            for target, target_doc, local_target in target_elements:
                target_point = target.Location.Point
                distance = calculate_distance(source_point, target_point)
                
                if distance < min_distance:
                    min_distance = distance
                    nearest_target = target
                    nearest_target_doc = target_doc
                    nearest_local_target = local_target
                
                pbar.update(1)
            
            results.append({
                'source': source,
                'source_doc': source_doc,
                'local_source': local_source,
                'nearest_target': nearest_target,
                'target_doc': nearest_target_doc,
                'local_target': nearest_local_target,
                'distance': min_distance
            })
    
    return results

def get_elements_by_category_and_description(category, search_strings, linked_docs, host_doc, workset_name=None):
    elements = []
    for linked_doc in linked_docs:
        collector = FilteredElementCollector(linked_doc).OfCategory(category).WhereElementIsNotElementType()
        for element in collector:
            element_type = linked_doc.GetElement(element.GetTypeId())
            if element_type:
                for param in element_type.Parameters:
                    if param.HasValue and param.StorageType == StorageType.String:
                        value = param.AsString().lower()
                        if any(search.lower() in value for search in search_strings):
                            # Try to get the local representation
                            local_element = None
                            try:
                                # First, try to find a matching element in the specified workset
                                if workset_name:
                                    workset_table = host_doc.GetWorksetTable()
                                    workset = next((ws for ws in workset_table.GetWorksets() if ws.Name == workset_name), None)
                                    if workset:
                                        local_collector = FilteredElementCollector(host_doc).OfCategory(category).WhereElementIsNotElementType().OnWorkset(workset.Id)
                                        for local_elem in local_collector:
                                            if local_elem.Name == element.Name and local_elem.GetTypeId().IntegerValue == element.GetTypeId().IntegerValue:
                                                local_element = local_elem
                                                break
                                
                                # If not found in the specific workset, try in all worksets
                                if local_element is None:
                                    local_collector = FilteredElementCollector(host_doc).OfCategory(category).WhereElementIsNotElementType()
                                    for local_elem in local_collector:
                                        if local_elem.Name == element.Name and local_elem.GetTypeId().IntegerValue == element.GetTypeId().IntegerValue:
                                            local_element = local_elem
                                            break
                            except Exception as e:
                                print(f"Error getting local element: {str(e)}")
                            
                            elements.append((element, linked_doc, local_element))
                            break
    print(f"Found {len(elements)} elements of category {category} matching search strings in linked documents")
    return elements
    
def print_element_info(elements, category_name):
    """
    Print detailed information about elements.
    
    Args:
        elements (List[Element]): List of elements to print information about.
        category_name (str): Name of the category for logging purposes.
    """
    print(f"\nDetailed information for {category_name}:")
    for i, element in enumerate(elements[:5], 1):  # Print info for first 5 elements
        print(f"  Element {i}:")
        print(f"    ID: {element.Id.IntegerValue}")
        print(f"    Category: {safe_get_property(element, 'Category')}")
        element_type = element.Document.GetElement(element.GetTypeId())
        print(f"    Family: {safe_get_property(element_type, 'Family')}")
        print(f"    Type: {safe_get_property(element_type, 'Name')}")
        for param_name in ["Family and Type Name", "Type Comments", "Description", "DESCRIPTION", "DESCRIPTION HEB", "Legend Description"]:
            param = element_type.LookupParameter(param_name)
            if param and param.HasValue:
                print(f"    {param_name}: {param.AsString()}")
        
        # Print all parameters and their values
        print("    All Parameters:")
        for param in element.Parameters:
            if param.HasValue:
                if param.StorageType == StorageType.String:
                    print(f"      {param.Definition.Name}: {param.AsString()}")
                elif param.StorageType == StorageType.Double:
                    print(f"      {param.Definition.Name}: {param.AsDouble()}")
                elif param.StorageType == StorageType.Integer:
                    print(f"      {param.Definition.Name}: {param.AsInteger()}")
                else:
                    print(f"      {param.Definition.Name}: <Unsupported StorageType>")
        
    print(f"  ... and {len(elements) - 5} more elements") if len(elements) > 5 else None

def create_excel_report(comparison_data, output_path):
    try:
        wb = openpyxl.Workbook()
        ws_details = wb.active
        ws_details.title = "Type Comparison Details"
        
        headers = ["Source Linked UniqueId", "Source Local UniqueId", "Source Document", "Source Category", "Source Family", "Source Type", "Source Mark",
                   "Nearest Target Linked UniqueId", "Nearest Target Local UniqueId", "Target Document", "Target Category", "Target Family", "Target Type", "Target Mark",
                   "Distance (m)"]
        
        for col, header in enumerate(headers, start=1):
            cell = ws_details.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        for row, data in enumerate(comparison_data, start=2):
            source = data['source']
            local_source = data['local_source']
            target = data['nearest_target']
            local_target = data['local_target']
            
            ws_details.cell(row=row, column=1, value=source.UniqueId)
            ws_details.cell(row=row, column=2, value=local_source.UniqueId if local_source else "N/A")
            ws_details.cell(row=row, column=3, value=data['source_doc'].Title)
            ws_details.cell(row=row, column=4, value=safe_get_property(source, 'Category'))
            ws_details.cell(row=row, column=5, value=safe_get_property(source, 'Family'))
            ws_details.cell(row=row, column=6, value=safe_get_property(source, 'Name'))
            ws_details.cell(row=row, column=7, value=get_parameter_value(source, 'Mark'))
            
            ws_details.cell(row=row, column=8, value=target.UniqueId if target else "N/A")
            ws_details.cell(row=row, column=9, value=local_target.UniqueId if local_target else "N/A")
            ws_details.cell(row=row, column=10, value=data['target_doc'].Title if target else "N/A")
            ws_details.cell(row=row, column=11, value=safe_get_property(target, 'Category') if target else "N/A")
            ws_details.cell(row=row, column=12, value=safe_get_property(target, 'Family') if target else "N/A")
            ws_details.cell(row=row, column=13, value=safe_get_property(target, 'Name') if target else "N/A")
            ws_details.cell(row=row, column=14, value=get_parameter_value(target, 'Mark') if target else "N/A")
            
            ws_details.cell(row=row, column=15, value=round(data['distance'], 2))

        # Adjust column widths and add filters
        for column in ws_details.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            column_letter = column[0].column_letter
            ws_details.column_dimensions[column_letter].width = max_length + 2

        ws_details.auto_filter.ref = ws_details.dimensions
        ws_details.freeze_panes = ws_details['A2']

        wb.save(output_path)
        print(f"Excel report saved to {output_path}")
    except Exception as e:
        print(f"Error creating Excel report to {output_path}: {str(e)}")
        print(traceback.format_exc())
        
def get_folder_path(prompt):
    try:
        from System.Windows.Forms import FolderBrowserDialog, DialogResult
        folder_dialog = FolderBrowserDialog()
        folder_dialog.Description = prompt
        if folder_dialog.ShowDialog() == DialogResult.OK:
            return folder_dialog.SelectedPath
    except:
        pass

    try:
        folder_path = input(f"{prompt} (enter full path): ")
        if os.path.isdir(folder_path):
            return folder_path
        else:
            print("Invalid directory. Please try again.")
    except:
        pass

    try:
        path = TaskDialog.Show("Folder Path Input",
                               prompt,
                               TaskDialogCommonButtons.Ok | TaskDialogCommonButtons.Cancel,
                               TaskDialogResult.Ok)
        if path == TaskDialogResult.Ok:
            folder_path = TaskDialog.Show("Folder Path Input", "Enter the full folder path:")
            if os.path.isdir(folder_path):
                return folder_path
            else:
                TaskDialog.Show("Error", "Invalid directory. Please try again.")
    except:
        pass

    TaskDialog.Show("Error", "Unable to get folder path input.")
    return None

def validate_folder_path(folder_path):
    if not folder_path:
        return False
    if not os.path.exists(folder_path):
        TaskDialog.Show("Error", f"The specified folder does not exist: {folder_path}")
        return False
    return True

def safe_get_property(element, property_name):
    """
    Safely get a property value from an element or its type.
    
    Args:
        element (Element): The Revit element.
        property_name (str): The name of the property to retrieve.
    
    Returns:
        str: The property value if accessible, or an error message if not.
    """
    try:
        if property_name == 'Category':
            return element.Category.Name if element.Category else "<Category not found>"
        elif property_name == 'Family':
            return element.Symbol.Family.Name if element.Symbol and element.Symbol.Family else "<Family not found>"
        elif hasattr(element, property_name):
            return str(getattr(element, property_name))
        elif hasattr(element.Symbol, property_name):
            return str(getattr(element.Symbol, property_name))
        else:
            return f"<{property_name} not found>"
    except Exception as e:
        return f"<Error accessing {property_name}: {str(e)}>"

def get_parameter_value(element, param_name):
    """
    Get the value of a parameter from an element.
    
    Args:
        element (Element): The Revit element.
        param_name (str): The name of the parameter.
    
    Returns:
        str: The parameter value if found, or None if not found.
    """
    param = element.LookupParameter(param_name)
    if param and param.HasValue:
        if param.StorageType == StorageType.String:
            return param.AsString()
        elif param.StorageType == StorageType.Double:
            return str(param.AsDouble())
        elif param.StorageType == StorageType.Integer:
            return str(param.AsInteger())
    return None
    
def main():
    clear_all_lru_caches()

    folder_path = get_folder_path("Select a folder to save the Excel report")
    if not validate_folder_path(folder_path):
        return

    linked_docs = get_linked_documents()
    if not linked_docs:
        TaskDialog.Show("Error", "No linked documents found.")
        return

    power_devices = get_elements_by_category_and_description(
        BuiltInCategory.OST_MechanicalEquipment,
        ["AES", "מעבים"],
        linked_docs,
        doc  # Pass the host document
    )
    outlets = get_elements_by_category_and_description(
        BuiltInCategory.OST_ElectricalFixtures,
        ["Socket", "בית תקע"],
        linked_docs,
        doc  # Pass the host document
    )

    results = calculate_nearest_distance(power_devices, outlets)

    print(f"Total results: {len(results)}")

    output_path = os.path.join(folder_path, "Comparison_Report.xlsx")
    create_excel_report(results, output_path)

    view_name = "VDC - HVAC ELEC verification"
    
    # Check if the view already exists
    existing_view = next((view for view in FilteredElementCollector(doc).OfClass(View3D) 
                          if not view.IsTemplate and view.Name == view_name), None)
    
    is_new_view = existing_view is None
    modified_view = create_or_modify_3d_view(doc, power_devices, outlets, view_name, is_new_view)
    
    if modified_view:
        uidoc.ActiveView = modified_view
        action = "created" if is_new_view else "modified"
        TaskDialog.Show("Success", f"3D view '{view_name}' {action} successfully.")
    else:
        action = "create" if is_new_view else "modify"
        TaskDialog.Show("Error", f"Failed to {action} 3D view '{view_name}'. Check the script output for details.")

    clear_all_lru_caches()

if __name__ == '__main__':
    main()
