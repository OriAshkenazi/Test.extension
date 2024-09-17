#! python3

import clr
import os
import sys
import traceback
import math
import logging
from collections import defaultdict
from typing import List, Tuple
from functools import wraps, lru_cache
from tqdm import tqdm

# Add references to Revit API
clr.AddReference('RevitAPI')
clr.AddReference('RevitAPIUI')
from Autodesk.Revit.DB import *
from Autodesk.Revit.UI import *
from Autodesk.Revit.Exceptions import InvalidOperationException

# Import .NET libraries for forms
clr.AddReference('System.Windows.Forms')
clr.AddReference('System.Drawing')
from System.Windows.Forms import (
    Form, Label, TextBox, Button, DialogResult, FolderBrowserDialog,
    MessageBox, MessageBoxButtons, MessageBoxIcon, FormStartPosition
)
from System.Drawing import Point, Size

import System
from System import Guid
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment

import numpy as np
from scipy.spatial import cKDTree

# Get the Revit application and document
uidoc = __revit__.ActiveUIDocument
doc = uidoc.Document

lru_cached_functions = []

def tracked_lru_cache(*args, **kwargs):
    """
    Decorator to track functions using LRU cache.

    Args:
        *args: Positional arguments for lru_cache.
        **kwargs: Keyword arguments for lru_cache.

    Returns:
        Decorator function to wrap the target function.
    """
    def decorator(func):
        cached_func = lru_cache(*args, **kwargs)(func)
        lru_cached_functions.append(cached_func)
        return cached_func
    return decorator

def clear_all_lru_caches():
    """
    Clear all LRU caches for tracked functions.
    """
    for func in lru_cached_functions:
        func.cache_clear()

def get_linked_documents():
    linked_docs = []
    link_instances = FilteredElementCollector(doc).OfClass(RevitLinkInstance)

    for link_instance in link_instances:
        linked_doc = link_instance.GetLinkDocument()
        if linked_doc is not None:
            linked_docs.append(linked_doc)
    logging.info(f"Found {len(linked_docs)} linked documents.")
    return linked_docs

def calculate_transformation(source_doc, dest_doc):
    """
    Calculate the transformation between a source document and a destination document.

    Args:
        source_doc (Document): The source Revit document.
        dest_doc (Document): The destination Revit document.

    Returns:
        Transform: The transformation from the source to the destination document.
    """
    link_instances = FilteredElementCollector(dest_doc).OfClass(RevitLinkInstance)
    source_link_instance = next((link for link in link_instances if link.GetLinkDocument().Title == source_doc.Title), None)

    if source_link_instance is None:
        raise Exception("Could not find the link instance in the destination document")

    return source_link_instance.GetTotalTransform()

def get_elements_by_category_and_description(category, search_strings, linked_docs, host_doc, workset_name=None):
    elements = []
    for linked_doc in linked_docs:
        collector = FilteredElementCollector(linked_doc).OfCategory(category).WhereElementIsNotElementType()
        for element in collector:
            element_type = linked_doc.GetElement(element.GetTypeId())
            if element_type:
                for param in element_type.Parameters:
                    if param.HasValue and param.StorageType == StorageType.String:
                        value = param.AsString().lower()
                        if any(search.lower() in value for search in search_strings):
                            # Try to get the local representation
                            local_element = None
                            try:
                                # Get transformation from linked_doc to host_doc
                                transform = calculate_transformation(linked_doc, host_doc)
                                # Transform the element's location point
                                if element.Location and isinstance(element.Location, LocationPoint):
                                    point = element.Location.Point
                                    transformed_point = transform.OfPoint(point)
                                    # Find the closest element in host_doc at the transformed location
                                    local_collector = FilteredElementCollector(host_doc).OfCategory(category).WhereElementIsNotElementType()
                                    for local_elem in local_collector:
                                        if local_elem.Location and isinstance(local_elem.Location, LocationPoint):
                                            local_point = local_elem.Location.Point
                                            distance = transformed_point.DistanceTo(local_point)
                                            if distance < 0.01:  # Threshold for matching elements
                                                local_element = local_elem
                                                break
                            except Exception as e:
                                pass  # If transformation fails, ignore and proceed

                            elements.append((element, linked_doc, local_element))
                            break
    category_name = Category.GetCategory(host_doc, category).Name
    logging.info(f"Found {len(elements)} elements of category '{category_name}' (ID: {category}) matching search strings in linked documents")
    return elements

@tracked_lru_cache(maxsize=None)
def calculate_distance(source_point, target_point):
    """
    Calculate the distance between two points and convert it to meters.

    Args:
        source_point (XYZ): The source point.
        target_point (XYZ): The target point.

    Returns:
        float: The distance between the source and target points in meters.
    """
    distance_feet = source_point.DistanceTo(target_point)
    return distance_feet * 0.3048  # Convert feet to meters

def calculate_top_n_nearest(source_elements, target_elements, N):
    """
    Calculate the top N nearest target elements for each source element.

    Args:
        source_elements (List[Tuple[Element, Document, Element]]): List of source elements.
        target_elements (List[Tuple[Element, Document, Element]]): List of target elements.
        N (int): Number of nearest targets to find.

    Returns:
        List[Dict]: List of dictionaries containing the source element and its top N nearest targets.
    """
    results = []

    # Prepare source points
    source_points = []
    source_data = []
    for source, source_doc, local_source in source_elements:
        if source.Location and isinstance(source.Location, LocationPoint):
            point = source.Location.Point
            # Transform the point to host document coordinates if necessary
            if source_doc != doc:
                transform = calculate_transformation(source_doc, doc)
                point = transform.OfPoint(point)
            source_points.append((point.X, point.Y, point.Z))
            source_data.append({'source': source, 'source_doc': source_doc, 'local_source': local_source})

    # Prepare target points
    target_points = []
    target_data = []
    for target, target_doc, local_target in target_elements:
        if target.Location and isinstance(target.Location, LocationPoint):
            point = target.Location.Point
            # Transform the point to host document coordinates if necessary
            if target_doc != doc:
                transform = calculate_transformation(target_doc, doc)
                point = transform.OfPoint(point)
            target_points.append((point.X, point.Y, point.Z))
            target_data.append({'target': target, 'target_doc': target_doc, 'local_target': local_target})

    if not target_points:
        logging.error("No target elements found with valid locations.")
        sys.exit()

    # Build k-d tree for target points
    tree = cKDTree(target_points)

    results = []

    # Use tqdm for the progress bar
    total_comparisons = len(source_points)
    for idx, source_point in enumerate(tqdm(source_points, desc="Calculating distances", unit="element", total=total_comparisons)):
        distances, indices = tree.query(source_point, k=N)
        if N == 1:
            distances = [distances]
            indices = [indices]
        top_n_targets = []
        for dist, idx_target in zip(distances, indices):
            target_info = target_data[idx_target]
            top_n_targets.append({
                'nearest_target': target_info['target'],
                'target_doc': target_info['target_doc'],
                'local_target': target_info['local_target'],
                'distance': dist * 0.3048  # Convert feet to meters
            })
        result = source_data[idx]
        result['top_n_targets'] = top_n_targets
        results.append(result)

    return results

def create_excel_report(comparison_data, output_path, N):
    """
    Create an Excel report of the comparison data.

    Args:
        comparison_data (List[Dict]): The comparison data.
        output_path (str): The path to save the Excel report.
        N (int): Number of nearest targets listed for each source.
    """
    try:
        wb = openpyxl.Workbook()
        ws_details = wb.active
        ws_details.title = "HVAC to Socket Distance"

        headers = [
            "Source UniqueId", "Source Document", "Source Category", "Source Family", "Source Type", "Source Mark",
        ]
        for n in range(1, N+1):
            headers.extend([
                f"Target {n} UniqueId", f"Target {n} Document", f"Target {n} Category",
                f"Target {n} Family", f"Target {n} Type", f"Target {n} Mark", f"Distance {n} (m)"
            ])

        for col, header in enumerate(headers, start=1):
            cell = ws_details.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        for row, data in enumerate(comparison_data, start=2):
            source = data['source']
            local_source = data['local_source']

            ws_details.cell(row=row, column=1, value=source.UniqueId)
            ws_details.cell(row=row, column=2, value=data['source_doc'].Title)
            ws_details.cell(row=row, column=3, value=safe_get_property(source, 'Category'))
            ws_details.cell(row=row, column=4, value=safe_get_property(source, 'Family'))
            ws_details.cell(row=row, column=5, value=safe_get_property(source, 'Name'))
            ws_details.cell(row=row, column=6, value=get_parameter_value(source, 'Mark'))

            col_offset = 7
            for idx, target_info in enumerate(data['top_n_targets'], start=1):
                target = target_info['nearest_target']
                ws_details.cell(row=row, column=col_offset, value=target.UniqueId if target else "N/A")
                ws_details.cell(row=row, column=col_offset+1, value=target_info['target_doc'].Title if target else "N/A")
                ws_details.cell(row=row, column=col_offset+2, value=safe_get_property(target, 'Category') if target else "N/A")
                ws_details.cell(row=row, column=col_offset+3, value=safe_get_property(target, 'Family') if target else "N/A")
                ws_details.cell(row=row, column=col_offset+4, value=safe_get_property(target, 'Name') if target else "N/A")
                ws_details.cell(row=row, column=col_offset+5, value=get_parameter_value(target, 'Mark') if target else "N/A")
                ws_details.cell(row=row, column=col_offset+6, value=round(target_info['distance'], 2))
                col_offset += 7

        # Adjust column widths and add filters
        for column in ws_details.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            column_letter = column[0].column_letter
            ws_details.column_dimensions[column_letter].width = max_length + 2

        ws_details.auto_filter.ref = ws_details.dimensions
        ws_details.freeze_panes = ws_details['A2']

        wb.save(output_path)
        logging.info(f"Excel report saved to {output_path}")
    except Exception as e:
        logging.error(f"Error creating Excel report to {output_path}: {str(e)}")
        logging.error(traceback.format_exc())

def safe_get_property(element, property_name):
    """
    Safely get a property value from an element or its type.

    Args:
        element (Element): The Revit element.
        property_name (str): The name of the property to retrieve.

    Returns:
        str: The property value if accessible, or an error message if not.
    """
    try:
        if property_name == 'Category':
            return element.Category.Name if element.Category else "<Category not found>"
        elif property_name == 'Family':
            symbol = element.Symbol if hasattr(element, 'Symbol') else None
            if symbol and symbol.Family:
                return symbol.Family.Name
            else:
                return "<Family not found>"
        elif hasattr(element, property_name):
            return str(getattr(element, property_name))
        elif hasattr(element.Symbol, property_name):
            return str(getattr(element.Symbol, property_name))
        else:
            return f"<{property_name} not found>"
    except Exception as e:
        return f"<Error accessing {property_name}: {str(e)}>"

def get_parameter_value(element, param_name):
    """
    Get the value of a parameter from an element.

    Args:
        element (Element): The Revit element.
        param_name (str): The name of the parameter.

    Returns:
        str: The parameter value if found, or None if not found.
    """
    param = element.LookupParameter(param_name)
    if param and param.HasValue:
        if param.StorageType == StorageType.String:
            return param.AsString()
        elif param.StorageType == StorageType.Double:
            return str(param.AsDouble())
        elif param.StorageType == StorageType.Integer:
            return str(param.AsInteger())
    return None

# Custom Form for User Input
class InputForm(Form):
    def __init__(self):
        self.Text = "Specify Parameters"
        self.Width = 500
        self.Height = 300
        self.StartPosition = FormStartPosition.CenterScreen

        label_width = 200
        textbox_width = 250
        control_height = 20
        vertical_space = 30
        left_margin = 10
        top_margin = 20

        # Number of Nearest Sockets (N)
        self.labelN = Label()
        self.labelN.Text = "Number of Nearest Sockets (N):"
        self.labelN.Size = Size(label_width, control_height)
        self.labelN.Location = Point(left_margin, top_margin)
        self.Controls.Add(self.labelN)

        self.textN = TextBox()
        self.textN.Text = "3"
        self.textN.Size = Size(textbox_width, control_height)
        self.textN.Location = Point(left_margin + label_width, top_margin)
        self.Controls.Add(self.textN)

        # Search Strings for HVAC Units
        self.labelHVAC = Label()
        self.labelHVAC.Text = "Search Strings for HVAC Units:"
        self.labelHVAC.Size = Size(label_width, control_height)
        self.labelHVAC.Location = Point(left_margin, top_margin + vertical_space)
        self.Controls.Add(self.labelHVAC)

        self.textHVAC = TextBox()
        self.textHVAC.Text = "AES, מעבים"
        self.textHVAC.Size = Size(textbox_width, control_height)
        self.textHVAC.Location = Point(left_margin + label_width, top_margin + vertical_space)
        self.Controls.Add(self.textHVAC)

        # Search Strings for Sockets
        self.labelSockets = Label()
        self.labelSockets.Text = "Search Strings for Sockets:"
        self.labelSockets.Size = Size(label_width, control_height)
        self.labelSockets.Location = Point(left_margin, top_margin + 2 * vertical_space)
        self.Controls.Add(self.labelSockets)

        self.textSockets = TextBox()
        self.textSockets.Text = "Socket, בית תקע"
        self.textSockets.Size = Size(textbox_width, control_height)
        self.textSockets.Location = Point(left_margin + label_width, top_margin + 2 * vertical_space)
        self.Controls.Add(self.textSockets)

        # Distance Threshold
        self.labelThreshold = Label()
        self.labelThreshold.Text = "Distance Threshold (meters):"
        self.labelThreshold.Size = Size(label_width, control_height)
        self.labelThreshold.Location = Point(left_margin, top_margin + 3 * vertical_space)
        self.Controls.Add(self.labelThreshold)

        self.textThreshold = TextBox()
        self.textThreshold.Text = "5"
        self.textThreshold.Size = Size(textbox_width, control_height)
        self.textThreshold.Location = Point(left_margin + label_width, top_margin + 3 * vertical_space)
        self.Controls.Add(self.textThreshold)

        # OK and Cancel buttons
        self.okButton = Button()
        self.okButton.Text = "OK"
        self.okButton.Size = Size(100, 30)
        self.okButton.Location = Point(100, top_margin + 4 * vertical_space + 10)
        self.okButton.DialogResult = DialogResult.OK
        self.Controls.Add(self.okButton)

        self.cancelButton = Button()
        self.cancelButton.Text = "Cancel"
        self.cancelButton.Size = Size(100, 30)
        self.cancelButton.Location = Point(250, top_margin + 4 * vertical_space + 10)
        self.cancelButton.DialogResult = DialogResult.Cancel
        self.Controls.Add(self.cancelButton)

        self.AcceptButton = self.okButton
        self.CancelButton = self.cancelButton

def main():
    clear_all_lru_caches()

    # Set up logging to output to console
    logging.basicConfig(level=logging.INFO, format='%(message)s')

    # Get user inputs using custom form
    form = InputForm()
    result = form.ShowDialog()

    if result == DialogResult.OK:
        try:
            N = int(form.textN.Text)
            hvac_search_strings = [s.strip() for s in form.textHVAC.Text.split(',')]
            socket_search_strings = [s.strip() for s in form.textSockets.Text.split(',')]
            distance_threshold = float(form.textThreshold.Text)
        except ValueError:
            MessageBox.Show("Invalid input values. Please ensure numeric values are entered where required.", "Input Error", MessageBoxButtons.OK, MessageBoxIcon.Error)
            return
    else:
        # User cancelled
        MessageBox.Show("The script was cancelled by the user.", "Script Cancelled", MessageBoxButtons.OK, MessageBoxIcon.Information)
        return

    # Get folder path using FolderBrowserDialog
    folder_dialog = FolderBrowserDialog()
    folder_dialog.Description = "Select a folder to save the Excel report"
    if folder_dialog.ShowDialog() == DialogResult.OK:
        folder_path = folder_dialog.SelectedPath
    else:
        MessageBox.Show("No folder selected. Script will exit.", "Script Cancelled", MessageBoxButtons.OK, MessageBoxIcon.Information)
        return

    if not folder_path or not os.path.exists(folder_path):
        MessageBox.Show("Invalid folder path selected. Script will exit.", "Error", MessageBoxButtons.OK, MessageBoxIcon.Error)
        return

    output_path = os.path.join(folder_path, "HVAC_Socket_Distance_Report.xlsx")

    # Get linked documents
    linked_docs = get_linked_documents()
    if not linked_docs:
        MessageBox.Show("No linked documents found.", "Error", MessageBoxButtons.OK, MessageBoxIcon.Error)
        return

    # Get elements
    logging.info("Collecting HVAC units...")
    power_devices = get_elements_by_category_and_description(
        BuiltInCategory.OST_MechanicalEquipment,
        hvac_search_strings,
        linked_docs,
        doc
    )

    logging.info("Collecting sockets...")
    sockets = get_elements_by_category_and_description(
        BuiltInCategory.OST_ElectricalFixtures,
        socket_search_strings,
        linked_docs,
        doc
    )

    if not power_devices:
        MessageBox.Show("No HVAC units found matching the search criteria.", "Error", MessageBoxButtons.OK, MessageBoxIcon.Error)
        return
    if not sockets:
        MessageBox.Show("No sockets found matching the search criteria.", "Error", MessageBoxButtons.OK, MessageBoxIcon.Error)
        return

    # Calculate distances
    logging.info("Calculating distances...")
    results = calculate_top_n_nearest(power_devices, sockets, N)

    logging.info(f"Total results: {len(results)}")

    # Create Excel report
    logging.info("Creating Excel report...")
    create_excel_report(results, output_path, N)

    logging.info(f"Excel report saved to {output_path}")

    logging.info("Script completed successfully.")

if __name__ == '__main__':
    main()
