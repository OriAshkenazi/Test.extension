#! python3

import clr
import functools
import traceback
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import datetime
import os
import logging
from pathlib import Path
from typing import Set, List, Dict, Tuple, Optional
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.views import SheetView, Pane
from shapely.geometry import Polygon, MultiPolygon, LineString, Point, GeometryCollection
from shapely.ops import unary_union
from shapely.validation import make_valid

clr.AddReference('RevitAPI')
clr.AddReference('System')
from Autodesk.Revit.DB import *
from Autodesk.Revit.Exceptions import InvalidOperationException

# Get the current document
doc = __revit__.ActiveUIDocument.Document

# List to keep track of all LRU cached functions
lru_cached_functions = []

def tracked_lru_cache(*args, **kwargs):
    def decorator(func):
        cached_func = functools.lru_cache(*args, **kwargs)(func)
        lru_cached_functions.append(cached_func)
        return cached_func
    return decorator

def setup_logging(prefix, timestamp):
    '''
    Setup logging to a file in the specified directory.

    Args:
        prefix (str): The prefix for the export directory.
        timestamp (str): The timestamp for the export directory.
    '''
    # Use the user's Documents folder as the base directory
    base_dir = Path.home() / "Documents"
    log_dir = base_dir / "Shapir" / "Exports" / prefix / timestamp
    
    # Create the directory
    log_dir.mkdir(parents=True, exist_ok=True)
    
    log_file = log_dir / "debug_log.txt"
    logging.basicConfig(filename=str(log_file), level=logging.DEBUG,
                        format='%(asctime)s - %(levelname)s - %(message)s')

    # Log the directory creation
    logging.info(f"Log directory created: {log_dir}")

def timing(func):
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        start_time = datetime.datetime.now()
        result = func(*args, **kwargs)
        current_time = datetime.datetime.now()
        elapsed_time = (current_time - start_time).total_seconds()
        
        # Assuming the first argument is the ceiling index and the second is the total number of ceilings
        i, total_ceilings = args[0], args[1]
        
        total_time_expected = elapsed_time / i * total_ceilings
        percent = (elapsed_time / total_time_expected) * 100 if total_time_expected > 0 else 0
        time_to_finish = total_time_expected - elapsed_time
        avg_processing_time = elapsed_time / i
        
        print(f"Time to finish: {abs(time_to_finish):.2f}s ({percent:.2f}%);   Elapsed time: {elapsed_time:.2f}s;    Avg. processing time: {avg_processing_time:.4f}s")
        
        return result
    return wrapper

@tracked_lru_cache(maxsize=None)
def get_element_geometry(element_id):
    """
    Memoized function to get the geometry of a Revit element.
    
    Args:
        element_id (ElementId): The ID of the Revit element.
    Returns:
        tuple or None: A tuple of Solid objects representing the element's geometry, or None if no valid geometry is found.
    """
    element = doc.GetElement(element_id)
    try:
        geom = element.get_Geometry(Options())
        if geom:
            solids = [solid for solid in geom if isinstance(solid, Solid) and solid.Volume > 0]
            return tuple(solids) if solids else None
        return None
    except Exception as e:
        logging.exception(f"Error in get_element_geometry: {e}")
        return None

@tracked_lru_cache(maxsize=None)
def get_element_bounding_box(element_id):
    """
    Memoized function to get the bounding box of an element.
    
    Args:
        element_id (ElementId): The ID of the Revit element.
    Returns:
        BoundingBoxXYZ: The bounding box of the element.
    """
    element = doc.GetElement(element_id)
    return element.get_BoundingBox(None)

@tracked_lru_cache(maxsize=None)
def project_to_xy_plane(solid, is_ceiling=False):
    """
    Project the solid geometry to the XY plane, creating a single polygon for ceilings.
    
    Args:
        solid (Solid): The solid geometry.
        is_ceiling (bool): Whether the solid represents a ceiling.
    Returns:
        Polygon or list of Polygon: The projected polygon(s) on the XY plane.
    """
    vertices = []
    for face in solid.Faces:
        try:
            face_normal = face.ComputeNormal(UV(0.5, 0.5))  # Compute normal at the center of the face
            if abs(face_normal.Z) > 0.001:  # Consider faces that are not vertical (allowing for some tolerance)
                edge_loops = face.GetEdgesAsCurveLoops()
                for loop in edge_loops:
                    for curve in loop:
                        for point in curve.Tessellate():
                            vertices.append((point.X, point.Y))
        except Exception as e:
            logging.exception(f"Error processing face in project_to_xy_plane: {e}\n{traceback.format_exc()}")
    
    if not vertices:
        return None

    if is_ceiling:
        # For ceilings, create a single polygon using the convex hull (which is a polygon that encloses all points)
        try:
            from scipy.spatial import ConvexHull
            hull = ConvexHull(vertices)
            hull_points = [vertices[i] for i in hull.vertices]
            polygon = Polygon(hull_points)
            return polygon if polygon.is_valid else make_valid(polygon)
        except Exception as e:
            logging.exception(f"Error creating ceiling polygon: {e}\n{traceback.format_exc()}")
            return None
    else:
        # For rooms, create multiple polygons if necessary
        try:
            polygon = Polygon(vertices)
            if not polygon.is_valid:
                polygon = make_valid(polygon)
            return [polygon]
        except Exception as e:
            logging.exception(f"Error creating room polygon: {e}\n{traceback.format_exc()}")
            return None

def calculate_intersection_areas(room_id, ceiling_id):
    """
    Calculate both 3D and 2D (XY projection) intersection areas between a room and a ceiling.
    
    Args:
        room_id (ElementId): The ID of the room element.
        ceiling_id (ElementId): The ID of the ceiling element.
    Returns:
        tuple: (intersection_area_3d, intersection_area_xy) in square meters.
    """
    room_geom = get_element_geometry(room_id)
    ceiling_geom = get_element_geometry(ceiling_id)
    intersection_area_3d = 0
    intersection_area_xy = 0
    is_complex_shape = False

    if not room_geom or not ceiling_geom:
        return intersection_area_3d, intersection_area_xy

    try:
        # Calculate XY projection intersection
        room_polygons = []
        ceiling_polygons = []
        for solid in room_geom:
            room_polygons.extend(project_to_xy_plane(solid, is_ceiling=False))
        for solid in ceiling_geom:
            ceiling_polygon = project_to_xy_plane(solid, is_ceiling=True)
            if ceiling_polygon:
                ceiling_polygons.append(ceiling_polygon)
        
        room_polygons = [p for p in room_polygons if p is not None and p.is_valid]
        ceiling_polygons = [p for p in ceiling_polygons if p is not None and p.is_valid]
        
        if room_polygons and ceiling_polygons:
            room_multi_poly = unary_union(room_polygons)
            ceiling_multi_poly = unary_union(ceiling_polygons)
            # Plot the room_polygons and ceiling_polygons geometries accurately on the xy plane
            plt.figure()
            for poly in room_polygons:
                if isinstance(poly, Polygon):
                    x, y = poly.exterior.xy
                    plt.plot(x, y, color='blue', linewidth=5)
                elif isinstance(poly, MultiPolygon):
                    for sub_poly in poly:
                        x, y = sub_poly.exterior.xy
                        plt.plot(x, y, color='blue', linewidth=5)
            for poly in ceiling_polygons:
                if isinstance(poly, Polygon):
                    x, y = poly.exterior.xy
                    plt.plot(x, y, color='red', linewidth=1)
                elif isinstance(poly, MultiPolygon):
                    for sub_poly in poly:
                        x, y = sub_poly.exterior.xy
                        plt.plot(x, y, color='red', linewidth=1)
            plt.axis('equal')
            plt.savefig(r"C:\Mac\Home\Documents\Shapir\Exports\room_ceiling_plot_xy.png")

            # Plot the room_polygons and ceiling_polygons geometries on the yz plane in a new subplot
            plt.figure()
            for poly in room_polygons:
                if isinstance(poly, Polygon):
                    x, y = poly.exterior.xy
                    plt.plot(y, x, color='blue', linewidth=5)
                elif isinstance(poly, MultiPolygon):
                    for sub_poly in poly:
                        x, y = sub_poly.exterior.xy
                        plt.plot(y, x, color='blue', linewidth=5)
            for poly in ceiling_polygons:
                if isinstance(poly, Polygon):
                    x, y = poly.exterior.xy
                    plt.plot(y, x, color='red', linewidth=1)
                elif isinstance(poly, MultiPolygon):
                    for sub_poly in poly:
                        x, y = sub_poly.exterior.xy
                        plt.plot(y, x, color='red', linewidth=1)
            plt.axis('equal')
            plt.savefig(r"C:\Mac\Home\Documents\Shapir\Exports\room_ceiling_plot_yz.png")
            
            if room_multi_poly.intersects(ceiling_multi_poly):
                intersection = room_multi_poly.intersection(ceiling_multi_poly)
                if isinstance(intersection, (Polygon, MultiPolygon)):
                    intersection_area_xy = intersection.area * 0.092903  # Convert to square meters

        # Calculate 3D intersection
        room_solids = [solid for solid in room_geom if isinstance(solid, Solid) and solid.Volume > 0]
        ceiling_solids = [solid for solid in ceiling_geom if isinstance(solid, Solid) and solid.Volume > 0]

        if room_solids and ceiling_solids:
            try:
                # Use the first solid from each collection for the boolean operation
                room_solid = room_solids[0]
                ceiling_solid = ceiling_solids[0]
                intersection_solid = BooleanOperationsUtils.ExecuteBooleanOperation(
                    room_solid, ceiling_solid, BooleanOperationsType.Intersect)
                if intersection_solid and intersection_solid.Volume > 0:
                    intersection_area_3d = intersection_solid.Volume * 0.092903  # Convert to square meters
            except InvalidOperationException as e:
                logging.exception(f"Boolean operation failed: {e}; Room ID: {room_id.IntegerValue}, Ceiling ID: {ceiling_id.IntegerValue}")
                # If boolean operation fails, use the XY projection area as an approximation
                intersection_area_3d = intersection_area_xy
        
        room_area = get_room_details(doc.GetElement(room_id))[6]
        ceiling_area = get_ceiling_details(doc.GetElement(ceiling_id))[2]

        # Check if intersection area is larger than room or ceiling area
        if intersection_area_3d > room_area * 1.05 or intersection_area_3d > ceiling_area * 1.3 or \
           intersection_area_xy > room_area * 1.05 or intersection_area_xy > ceiling_area * 1.3 or \
           (intersection_area_3d > 0 and intersection_area_xy == 0):
            is_complex_shape = True

    except Exception as e:
        logging.exception(f"Error in calculate_intersection_areas: {e}\n{traceback.format_exc()}")

    return intersection_area_3d, intersection_area_xy, is_complex_shape

def save_and_close_figure(fig, output_path):
    '''
    Save the figure to the specified output path and close it.

    Args:
        fig (Figure): The figure to save.
        output_path (str): The path to save the figure.
    '''
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    fig.savefig(output_path)
    plt.close(fig)

def plot_room_ceiling(room_id, ceiling_id, output_path, ax1, ax2, prefix, timestamp):
    '''
    Plot the room and ceiling geometries on the specified ax1/ax2 plane with additional details.
    Args:
        room_id (ElementId): The ID of the room element.
        ceiling_id (ElementId): The ID of the ceiling element.
        output_path (str): The path to save the plot.
        ax1 (str): The first axis to plot (X, Y, or Z).
        ax2 (str): The second axis to plot (X, Y, or Z).
    '''
    fig = plt.figure(figsize=(12, 8))
    room_geom = get_element_geometry(room_id)
    ceiling_geom = get_element_geometry(ceiling_id)

    room_details = get_room_details(doc.GetElement(room_id))
    ceiling_details = get_ceiling_details(doc.GetElement(ceiling_id))

    def get_point_coords(point):
        coords = {'X': point.X, 'Y': point.Y, 'Z': point.Z}
        return coords[ax1], coords[ax2]

    for geom, color, linewidth in [(room_geom, 'blue', 5), (ceiling_geom, 'red', 1)]:
        for solid in geom:
            for face in solid.Faces:
                for loop in face.EdgeLoops:
                    for edge in loop:
                        p1 = edge.AsCurve().GetEndPoint(0)
                        p2 = edge.AsCurve().GetEndPoint(1)
                        x1, y1 = get_point_coords(p1)
                        x2, y2 = get_point_coords(p2)
                        fig.plot([x1, x2], [y1, y2], color=color, linewidth=linewidth)

    plt.axis('equal')
    
    # Add title and labels
    plt.title(f"Room {room_details[2]} - {room_details[1]} (Level: {room_details[3]})")
    plt.xlabel(f"{ax1} (meters)")
    plt.ylabel(f"{ax2} (meters)")

    # Add room information
    room_info = (
        f"Room Area: {room_details[6]:.2f} m²\n"
        f"Building: {room_details[4]}\n"
        f"Ceiling Finish: {room_details[5]}"
    )
    plt.text(0.02, 0.98, room_info, transform=ax.transAxes, verticalalignment='top', fontsize=8)

    # Add ceiling information
    ceiling_info = (
        f"Ceiling Type: {ceiling_details[1]}\n"
        f"Ceiling Area: {ceiling_details[2]:.2f} m²\n"
        f"Ceiling Level: {ceiling_details[3]}"
    )
    plt.text(0.02, 0.88, ceiling_info, transform=ax.transAxes, verticalalignment='top', fontsize=8)

    # Add legend
    ax.legend(['Room', 'Ceiling'], loc='lower right')

    output_plot_path = os.path.join(output_path, prefix, timestamp, "Plots", f"{ceiling_id}_{room_id}_{ax1}{ax2}.png")
    save_and_close_figure(fig, output_plot_path)
    logging.debug(f"Plot saved to: {output_plot_path}")

    logging.exception(f"Plot saved to: {output_plot_path}")

def check_direct_intersection(room_id, ceiling_id, prefix, timestamp):
    """
    Check if there's a direct intersection between room and ceiling geometries.
    
    Args:
        room_id (ElementId): The ID of the room element.
        ceiling_id (ElementId): The ID of the ceiling element.
    Returns:
        bool: True if there's a direct intersection, False otherwise.
    """
    room_geom = get_element_geometry(room_id)
    ceiling_geom = get_element_geometry(ceiling_id)

    # Plot the room and ceiling geometries accuratly on the xy plane
    # plot_room_ceiling(room_id, ceiling_id, r"C:\Mac\Home\Documents\Shapir\Exports\room_ceiling_plot_xy", 'X', 'Y', prefix, timestamp)

    # plot the room and ceiling geometries accuratly on the yz plane
    # plot_room_ceiling(room_id, ceiling_id, r"C:\Mac\Home\Documents\Shapir\Exports\room_ceiling_plot_yz", 'Y', 'Z', prefix, timestamp)

    for room_solid in room_geom:
        for ceiling_solid in ceiling_geom:
            try:
                intersection_solid = BooleanOperationsUtils.ExecuteBooleanOperation(
                    room_solid, ceiling_solid, BooleanOperationsType.Intersect)
                # print(f"Intersection volume: {intersection_solid.Volume}")
                # print(f"Room volume: {room_solid.Volume}")
                # print(f"Ceiling volume: {ceiling_solid.Volume}")
                if intersection_solid.Volume > 0:
                    return True
            except InvalidOperationException:
                # If Boolean operation fails, fall back to bounding box check
                room_bb = room_solid.GetBoundingBox()
                ceiling_bb = ceiling_solid.GetBoundingBox()
                if check_bounding_box_intersection(room_bb, ceiling_bb):
                    return True
    return False

def check_bounding_box_intersection(bb1, bb2):
    """
    Check if two bounding boxes intersect.
    
    Args:
        bb1 (BoundingBoxXYZ): First bounding box
        bb2 (BoundingBoxXYZ): Second bounding box
    Returns:
        bool: True if the bounding boxes intersect, False otherwise.
    """
    return (bb1.Min.X <= bb2.Max.X and bb1.Max.X >= bb2.Min.X and
            bb1.Min.Y <= bb2.Max.Y and bb1.Max.Y >= bb2.Min.Y and
            bb1.Min.Z <= bb2.Max.Z and bb1.Max.Z >= bb2.Min.Z)

def project_and_check_xy_intersection(room_id, ceiling_id):
    """
    Project geometries to XY plane and check for intersection.
    
    Args:
        room_id (ElementId): The ID of the room element.
        ceiling_id (ElementId): The ID of the ceiling element.
    Returns:
        bool: True if there's an intersection in the XY projection, False otherwise.
    """
    room_geom = get_element_geometry(room_id)
    ceiling_geom = get_element_geometry(ceiling_id)
    
    if not room_geom or not ceiling_geom:
        return False
    
    room_polygons = []
    ceiling_polygons = []
    
    for solid in room_geom:
        room_polygons.extend(project_to_xy_plane(solid, is_ceiling=False))
    for solid in ceiling_geom:
        ceiling_polygon = project_to_xy_plane(solid, is_ceiling=True)
        if ceiling_polygon:
            ceiling_polygons.append(ceiling_polygon)
    
    room_polygons = [p for p in room_polygons if p is not None and p.is_valid]
    ceiling_polygons = [p for p in ceiling_polygons if p is not None and p.is_valid]

    if not room_polygons or not ceiling_polygons:
        return False
    
    try:
        room_union = unary_union(room_polygons)
        ceiling_union = unary_union(ceiling_polygons)
        return room_union.intersects(ceiling_union)
    except Exception as e:
        logging.exception(f"Error in project_and_check_xy_intersection: {e}")
        logging.exception(traceback.format_exc())
        # Fallback to bounding box check
        room_bb = get_element_bounding_box(room_id)
        ceiling_bb = get_element_bounding_box(ceiling_id)
        return check_bounding_box_intersection(room_bb, ceiling_bb)

def delta_ceiling_above_room(room_id, ceiling_id):
    """
    Check if the ceiling is above the room.
    
    Args:
        room_id (ElementId): The ID of the room element.
        ceiling_id (ElementId): The ID of the ceiling element.
    Returns:
        int: distanse between the ceiling and the room.
    """
    room_bb = get_element_bounding_box(room_id)
    ceiling_bb = get_element_bounding_box(ceiling_id)
    # print(f"Delta: {ceiling_bb.Min.Z - room_bb.Max.Z}")
    return ceiling_bb.Min.Z - room_bb.Max.Z

@tracked_lru_cache(maxsize=None)
def get_room_details(room):
    """
    Get details of a room.
    
    Args:
        room (SpatialElement): The room element.
    Returns:
        Tuple[int, str, str, str, str, str, float]: The room details (room_id, room_name, room_number, room_level, room_building, room_ceiling_finish, room_area).
    """
    room_id = room.Id.IntegerValue
    room_name = room.get_Parameter(BuiltInParameter.ROOM_NAME).AsString()
    room_number = room.get_Parameter(BuiltInParameter.ROOM_NUMBER).AsString()
    room_level = doc.GetElement(room.LevelId).Name
    room_building_param = room.LookupParameter("בניין")
    room_building = room_building_param.AsString() if room_building_param else None
    room_ceiling_finish_param = room.LookupParameter("Room_שם גמר תקרה")
    room_ceiling_finish = room_ceiling_finish_param.AsString() if room_ceiling_finish_param else None
    room_area_param = room.LookupParameter("Area")
    room_area = room_area_param.AsDouble() * 0.092903 if room_area_param else None # Convert from square feet to square meters
    return room_id, room_name, room_number, room_level, room_building, room_ceiling_finish, room_area

@tracked_lru_cache(maxsize=None)
def get_ceiling_details(ceiling):
    """
    Get details of a ceiling.
    
    Args:
        ceiling (Ceiling): The ceiling element.
    Returns:
        Tuple[int, str, str, float, str]: The ceiling details (ceiling_id, ceiling_description, ceiling_area, ceiling_level).
    """
    ceiling_id = ceiling.Id.IntegerValue
    ceiling_type_element = doc.GetElement(ceiling.GetTypeId())
    
    ceiling_description_param = ceiling_type_element.LookupParameter("Description")
    ceiling_description = ceiling_description_param.AsString() if ceiling_description_param else None
    
    ceiling_area_param = ceiling.LookupParameter("Area")
    ceiling_area = ceiling_area_param.AsDouble() * 0.092903 if ceiling_area_param else None # Convert from square feet to square meters

    ceiling_level = doc.GetElement(ceiling.LevelId).Name if ceiling.LevelId else None

    return ceiling_id, ceiling_description, ceiling_area, ceiling_level

def custom_sort_key(value):
    """
    Custom sorting function to handle numeric strings as numbers and None values.

    Args:
        value: The value to be sorted.
    Returns:
        float: The value converted to a float, or infinity if the value is None.
    """
    if value is None:
        return float('inf')  # This will put None values at the end of the sort order
    try:
        return float(value)
    except ValueError:
        return value

def calculate_max_level_heights(doc, building_levels):
    '''
    Calculate the maximum level height for each building in the document.
    
    Args:
        doc (Document): The Revit document.
        building_levels (list): A list of sets, where each set contains the levels for a building.
    Returns:
        list: A list of maximum level heights, one for each building.
    '''
    max_heights = []
    
    for levels_set in building_levels:
        levels = FilteredElementCollector(doc).OfClass(Level).ToElements()
        levels = [level for level in levels if level.Id in levels_set]
        levels = sorted(levels, key=lambda l: l.Elevation)
        
        if len(levels) < 2:
            max_heights.append(float('inf'))  # If there's only one level or no levels, set max height to inf
            continue
        
        max_height = 0
        for i in range(1, len(levels)):
            height = levels[i].Elevation - levels[i-1].Elevation
            max_height = max(max_height, height)

        max_height = max_height * 0.3048  # Convert from feet to meters
        
        max_heights.append(max_height)
    
    # print(f"Max level heights: {max_heights}")

    return max_heights

def process_room(ceiling: Ceiling, room: SpatialElement, building_levels: List[set], prefix: str, timestamp: str) -> Tuple[List[Dict], List[Dict], List[ElementId], List[Dict], List[ElementId]]:
    '''
    Process a room and its relationship with a ceiling.

    Args:
        ceiling (Ceiling): The ceiling element to process.
        room (SpatialElement): The room element to process.
        building_levels (List[set]): List of sets containing building level ElementIds.
        prefix (str): Prefix for output files.
        timestamp (str): Timestamp for output files.
    Returns:
        Tuple containing:
        - List of relationship dictionaries
        - List of complex shape room dictionaries
        - List of room ElementIds without geometry
        - List of XY intersection dictionaries
        - List of room ElementIds with direct intersections
    '''
    relationships = []
    complex_shape_rooms = []
    no_geometry_rooms = []
    xy_intersections = []
    direct_intersections = []

    room_id = room.Id
    ceiling_id = ceiling.Id

    try:
        room_details = get_room_details(room)
        ceiling_details = get_ceiling_details(ceiling)
    except Exception as e:
        logging.exception(f"Error getting room or ceiling details: {e}")
        return relationships, complex_shape_rooms, no_geometry_rooms, xy_intersections, direct_intersections

    # Skip rooms without valid geometry
    if not get_element_geometry(room_id):
        no_geometry_rooms.append(room_id)
        return relationships, complex_shape_rooms, no_geometry_rooms, xy_intersections, direct_intersections

    intersection_data = check_intersections(room_id, ceiling_id, prefix, timestamp)
    if intersection_data is None:
        return relationships, complex_shape_rooms, no_geometry_rooms, xy_intersections, direct_intersections

    direct_intersection, xy_intersection, intersection_area_3d, intersection_area_xy, is_complex_shape, distance = intersection_data

    if direct_intersection or xy_intersection:
        room_data = create_room_data(room_details, ceiling_details, intersection_area_3d, intersection_area_xy, direct_intersection, xy_intersection, distance)

        if is_complex_shape:
            complex_shape_rooms.append(room_data)
        elif direct_intersection:
            direct_intersections.append(room_id)
            relationships.append(room_data)
        elif xy_intersection:
            xy_intersections.append(room_data)

    return relationships, complex_shape_rooms, no_geometry_rooms, xy_intersections, direct_intersections

def check_intersections(room_id: ElementId, ceiling_id: ElementId, prefix: str, timestamp: str) -> Optional[Tuple[bool, bool, float, float, bool, float]]:
    """Check for direct and XY projection intersections between a room and a ceiling."""
    try:
        direct_intersection = check_direct_intersection(room_id, ceiling_id, prefix, timestamp)
        xy_intersection = project_and_check_xy_intersection(room_id, ceiling_id) if not direct_intersection else True
        intersection_area_3d, intersection_area_xy, is_complex_shape = calculate_intersection_areas(room_id, ceiling_id)
        distance = delta_ceiling_above_room(room_id, ceiling_id)
        return direct_intersection, xy_intersection, intersection_area_3d, intersection_area_xy, is_complex_shape, distance
    except Exception as e:
        logging.exception(f"Error in intersection checks: {e}")
        return None

def create_room_data(room_details: Tuple, ceiling_details: Tuple, intersection_area_3d: float, intersection_area_xy: float, direct_intersection: bool, xy_intersection: bool, distance: float) -> Dict:
    """Create a dictionary with room and ceiling data."""
    return {
        'Room_ID': room_details[0],
        'Room_Name': room_details[1],
        'Room_Number': room_details[2],
        'Room_Level': room_details[3],
        'Room_Building': room_details[4],
        'Room_Area_sqm': room_details[6],
        'Room_Ceiling_Finish': room_details[5],
        'Ceiling_ID': ceiling_details[0],
        'Ceiling_Level': ceiling_details[3],
        'Ceiling_Description': ceiling_details[1],
        'Ceiling_Description_Old': ceiling_details[1],
        'Ceiling_Area_sqm': ceiling_details[2],
        'Intersection_Area_3D_sqm': intersection_area_3d,
        'Intersection_Area_XY_sqm': intersection_area_xy,
        'Intersection_Area_sqm': max(intersection_area_3d, intersection_area_xy),
        'Direct_Intersection': direct_intersection,
        'XY_Projection_Intersection': xy_intersection,
        'Distance': distance
    }

def check_ceiling_geometry(doc: Document, ceiling_id: ElementId) -> bool:
    '''
    Check if the ceiling has valid geometry.

    Args:
        doc (Document): The Revit document.
        ceiling_id (ElementId): The ID of the ceiling element.
    Returns:
        bool: True if the ceiling has valid geometry, False otherwise.
    '''
    if not get_element_geometry(ceiling_id):
        logging.exception(f"No geometry found for ceiling ID: {ceiling_id.IntegerValue}")
        return False
    return True

def process_room_intersections(doc: Document, room: SpatialElement, ceiling_id: ElementId, prefix: str, timestamp: str) -> Tuple[Optional[Dict], Optional[Dict], bool, bool]:
    '''
    Process the intersections between a room and a ceiling.

    Args:
        doc (Document): The Revit document.
        room (SpatialElement): The room element.
        ceiling_id (ElementId): The ID of the ceiling element.
        prefix (str): Prefix for output files.
        timestamp (str): Timestamp for output files.
    Returns:
        Tuple (list, list, bool, bool): room_data, complex_shape_room, is_direct, is_xy
    '''
    room_id = room.Id
    room_details = get_room_details(room)
    
    if not get_element_geometry(room_id):
        return None, None, False, False
    
    direct_intersection = check_direct_intersection(room_id, ceiling_id, prefix, timestamp)
    xy_intersection = project_and_check_xy_intersection(room_id, ceiling_id) if not direct_intersection else True
    
    if direct_intersection or xy_intersection:
        intersection_area_3d, intersection_area_xy, is_complex_shape = calculate_intersection_areas(room_id, ceiling_id)
        distance = delta_ceiling_above_room(room_id, ceiling_id)
        
        room_data = create_room_data(room_details, ceiling_details, intersection_area_3d, intersection_area_xy, direct_intersection, xy_intersection, distance)
        
        if is_complex_shape:
            return None, room_data, False, False
        else:
            return room_data, None, direct_intersection, xy_intersection
    
    return None, None, False, False

def handle_xy_intersections(xy_intersections: List[Dict], building_levels: List[set], doc: Document) -> List[Dict]:
    '''
    Handle intersections that are only in the XY projection.

    Args:
        xy_intersections (List[Dict]): List of room data dictionaries with only XY intersections.
        building_levels (List[set]): List of sets containing building level ElementIds.
        doc (Document): The Revit document.
    Returns:
        List[Dict]: List of room data dictionaries with leveled rooms.
    '''
    positive_distance_rooms = [room for room in xy_intersections if room['Distance'] >= 0]
    if not positive_distance_rooms:
        return []
    
    level_distances = calculate_max_level_heights(doc, building_levels)
    min_distance_room = min(positive_distance_rooms, key=lambda x: x['Distance'])
    min_distance_level = doc.GetElement(ElementId(min_distance_room['Room_ID'])).LevelId
    
    filtered_rooms = [
        room for room in positive_distance_rooms
        if doc.GetElement(ElementId(room['Room_ID'])).LevelId == min_distance_level
    ]
    
    return filter_leveled_rooms(filtered_rooms, level_distances)

def filter_leveled_rooms(filtered_rooms: List[Dict], level_distances: List[float]) -> List[Dict]:
    '''
    Filter the rooms based on the building levels and distances.

    Args:
        filtered_rooms (List[Dict]): List of room data dictionaries.
        level_distances (List[float]): List of maximum level heights for each building.
    Returns:
        List[Dict]: List of room data dictionaries with leveled rooms.
    '''
    leveled_rooms = []
    for room in filtered_rooms:
        if room['Room_Building'] == "A" and room['Distance'] <= level_distances[0]:
            leveled_rooms.append(room)
        elif room['Room_Building'] == "B" and room['Distance'] <= level_distances[1]:
            leveled_rooms.append(room)
        elif room['Distance'] <= level_distances[2]:
            leveled_rooms.append(room)
    return leveled_rooms

@timing
def process_ceiling(ceiling_index: int, total_ceilings: int, ceiling: Ceiling, room_elements: List[SpatialElement], building_levels: List[set], doc: Document, prefix: str, timestamp: str) -> Tuple[List[Dict], List[Tuple], List[Dict], List[ElementId]]:
    '''
    Process a ceiling and its relationships with rooms.

    Args:
        ceiling_index (int): The index of the ceiling in the list.
        total_ceilings (int): The total number of ceilings.
        ceiling (Ceiling): The ceiling element to process.
        room_elements (List[SpatialElement]): List of room elements.
        building_levels (List[set]): List of sets containing building level ElementIds.
        doc (Document): The Revit document.
        prefix (str): Prefix for output files.
        timestamp (str): Timestamp for output files.
    Returns:
        Tuple (list, list, list, list): relationships, ceiling_details, complex_shape_rooms, no_geometry_rooms
    '''
    ceiling_id = ceiling.Id
    ceiling_details = get_ceiling_details(ceiling)
    
    relationships = []
    complex_shape_rooms = []
    no_geometry_rooms = []
    
    if not check_ceiling_geometry(doc, ceiling_id):
        return [], [ceiling_details], [], [ceiling_id]
    
    direct_intersections = []
    xy_intersections = []
    
    for room in room_elements:
        room_data, complex_shape_room, is_direct, is_xy = process_room_intersections(doc, room, ceiling_id, prefix, timestamp)
        
        if complex_shape_room:
            complex_shape_rooms.append(complex_shape_room)
        elif room_data:
            if is_direct:
                direct_intersections.append(room.Id)
                relationships.append(room_data)
            elif is_xy:
                xy_intersections.append(room_data)
        elif not get_element_geometry(room.Id):
            no_geometry_rooms.append(room.Id)
    
    if not direct_intersections and xy_intersections:
        leveled_rooms = handle_xy_intersections(xy_intersections, building_levels, doc)
        relationships.extend(leveled_rooms)
    
    if not relationships:
        return [], [ceiling_details], complex_shape_rooms, no_geometry_rooms
    
    return relationships, [], complex_shape_rooms, no_geometry_rooms

def update_ceiling_description(df_relationships):
    if not df_relationships.empty:
        df_relationships.loc[(df_relationships['Ceiling_Description'] == '"פלב"מ "ריגיד') & 
                             (df_relationships['Intersection_Area_sqm'] > 0) &
                             (df_relationships['Intersection_Area_sqm'] / df_relationships['Ceiling_Area_sqm'] > 0.09) &
                             (df_relationships['Room_Ceiling_Finish'] == 'פנל מבודד צבוע'), 'Ceiling_Description'
        ] = 'פנל מבודד צבוע'
        df_relationships.loc[(df_relationships['Ceiling_Description'] == "תקרת פח מחורר/אקופון 60/60 אדוונטיג' בשילוב סינורי גבס") &
                             (df_relationships['Intersection_Area_sqm'] > 0) &
                             (df_relationships['Intersection_Area_sqm'] / df_relationships['Ceiling_Area_sqm'] > 0.09) &
                             (df_relationships['Room_Ceiling_Finish'] == 'פורניר  ע"ג MDF  גודל 60/60 בשילוב סינורי גבס'), 'Ceiling_Description'
        ] = 'פורניר  ע"ג MDF  גודל 60/60 בשילוב סינורי גבס'
        df_relationships.loc[(df_relationships['Ceiling_Description'] == "תקרת פח מחורר/אקופון 60/60 אדוונטיג' בשילוב סינורי גבס") &
                             (df_relationships['Intersection_Area_sqm'] > 0) &
                             (df_relationships['Intersection_Area_sqm'] / df_relationships['Ceiling_Area_sqm'] > 0.09) &
                             (df_relationships['Room_Ceiling_Finish'] == 'תקרת פח מחורר 60/60 בשילוב סינורי גבס'), 'Ceiling_Description'
        ] = 'תקרת פח מחורר 60/60 בשילוב סינורי גבס'
        return df_relationships
    else:
        logging.exception("Dataframe 'df_relationships' is empty")
        return df_relationships

def initialize_data_structures() -> Tuple[List, List, List, List, List]:
    '''
    Initialize the data structures for processing ceilings.

    Returns:
        Tuple containing:
        - List of relationship dictionaries
        - List of unrelated ceiling tuples
        - List of complex shape room dictionaries
        - List of room ElementIds without geometry
        - List of ceiling ElementIds without geometry
    '''
    return [], [], [], [], []

def process_all_ceilings(
    ceiling_elements: List[Ceiling],
    room_elements: List[SpatialElement],
    building_levels: List[Set[ElementId]],
    doc: Document,
    prefix: str,
    timestamp: str
) -> Tuple[List[Dict], List[Tuple], List[Dict], List[ElementId], List[ElementId]]:
    '''
    Process all ceilings and their relationships with rooms.

    Args:
        ceiling_elements (List[Ceiling]): List of ceiling elements.
        room_elements (List[SpatialElement]): List of room elements.
        building_levels (List[Set[ElementId]]): List of sets containing building level ElementIds.
        doc (Document): The Revit document.
        prefix (str): Prefix for output files.
        timestamp (str): Timestamp for output files.
    Returns:
        Tuple containing:
        - List of relationship dictionaries
        - List of unrelated ceiling tuples
        - List of complex shape room dictionaries
        - List of room ElementIds without geometry
        - List of ceiling ElementIds without geometry
    '''
    relationships, unrelated_ceilings, complex_shape_rooms, no_geometry_rooms, no_geometry_ceilings = initialize_data_structures()

    for i, ceiling in enumerate(ceiling_elements, 1):
        ceiling_results = process_ceiling(i, len(ceiling_elements), ceiling, room_elements, building_levels, doc, prefix, timestamp)
        
        relationships.extend(ceiling_results[0])
        unrelated_ceilings.extend(ceiling_results[1])
        complex_shape_rooms.extend(ceiling_results[2])
        no_geometry_rooms.extend(ceiling_results[3])
        if ceiling_results[1]:  # If ceiling is unrelated, it has no geometry
            no_geometry_ceilings.append(ceiling.Id)

    return relationships, unrelated_ceilings, complex_shape_rooms, no_geometry_rooms, no_geometry_ceilings

def handle_no_geometry_rooms(no_geometry_rooms: List[ElementId], doc: Document):
    '''
    Handle rooms without geometry by logging their details once.

    Args:
        no_geometry_rooms (List[ElementId]): List of room ElementIds without geometry.
        doc (Document): The Revit document.
    '''
    for room_id in set(no_geometry_rooms):  # Use set to remove duplicates
        room = doc.GetElement(room_id)
        room_name = room.get_Parameter(BuiltInParameter.ROOM_NAME).AsString()
        room_number = room.get_Parameter(BuiltInParameter.ROOM_NUMBER).AsString()
        room_building_param = room.LookupParameter("בניין")
        room_building = room_building_param.AsString() if room_building_param else None
        room_level = doc.GetElement(room.LevelId).Name
        logging.exception(f"No geometry found for room ID: {room_id.IntegerValue} - Building: {room_building}, Level: {room_level}, Name: {room_name}, Number: {room_number}")

def create_dataframes(
    relationships: List[Dict],
    unrelated_ceilings: List[Tuple],
    complex_shape_rooms: List[Dict]
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    '''
    Create DataFrames from the processed data.

    Args:
        relationships (List[Dict]): List of relationship dictionaries.
        unrelated_ceilings (List[Tuple]): List of unrelated ceiling tuples.
        complex_shape_rooms (List[Dict]): List of complex shape room dictionaries.
    Returns:
        Tuple of DataFrames for relationships, unrelated ceilings, and complex shape rooms.
    '''
    df_relationships = pd.DataFrame(relationships)
    df_unrelated = pd.DataFrame(unrelated_ceilings, columns=['Ceiling_ID', 'Ceiling_Description', 'Ceiling_Area_sqm', 'Ceiling_Level'])
    df_complex_shape_rooms = pd.DataFrame(complex_shape_rooms)
    return df_relationships, df_unrelated, df_complex_shape_rooms

def sort_and_update_dataframes(
    df_relationships: pd.DataFrame,
    df_unrelated: pd.DataFrame,
    df_complex_shape_rooms: pd.DataFrame
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    '''
    Sort and update the DataFrames.

    Args:
        df_relationships (pd.DataFrame): DataFrame for relationships.
        df_unrelated (pd.DataFrame): DataFrame for unrelated ceilings.
        df_complex_shape_rooms (pd.DataFrame): DataFrame for complex shape rooms.
    Returns:
        Tuple of sorted and updated DataFrames.
    '''
    if not df_relationships.empty:
        df_relationships = df_relationships.sort_values(
            by=['Room_Building', 'Room_Level', 'Room_Number', 'Ceiling_Description'],
            key=lambda x: x.map(custom_sort_key)
        )
        df_relationships = update_ceiling_description(df_relationships)

    if not df_unrelated.empty:
        df_unrelated = df_unrelated.sort_values(
            by=['Ceiling_Level', 'Ceiling_Description'],
            key=lambda x: x.map(custom_sort_key)
        )

    if not df_complex_shape_rooms.empty:
        df_complex_shape_rooms = df_complex_shape_rooms.sort_values(
            by=['Room_Building', 'Room_Level', 'Room_Number'],
            key=lambda x: x.map(custom_sort_key)
        )

    return df_relationships, df_unrelated, df_complex_shape_rooms

def find_ceiling_room_relationships(
    room_elements: List[SpatialElement],
    ceiling_elements: List[Ceiling],
    building_A_levels: Set[ElementId],
    building_B_levels: Set[ElementId],
    common_levels: Set[ElementId],
    doc: Document,
    prefix: str,
    timestamp: str
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    '''
    Find the relationship between ceilings and rooms.

    Args:
        room_elements: List of room elements.
        ceiling_elements: List of ceiling elements.
        building_A_levels: Set of level IDs for building A.
        building_B_levels: Set of level IDs for building B.
        common_levels: Set of level IDs for common levels.
        doc: The current Revit document.
        prefix: The prefix for the output files.
        timestamp: The timestamp for the output files.
    Returns:
        Tuple of DataFrames for relationships, unrelated ceilings, and complex shape rooms.
    '''
    try:
        building_levels = [building_A_levels, building_B_levels, common_levels]

        relationships, unrelated_ceilings, complex_shape_rooms, no_geometry_rooms, no_geometry_ceilings = process_all_ceilings(
            ceiling_elements, room_elements, building_levels, doc, prefix, timestamp
        )

        handle_no_geometry_rooms(no_geometry_rooms, doc)

        df_relationships, df_unrelated, df_complex_shape_rooms = create_dataframes(
            relationships, unrelated_ceilings, complex_shape_rooms
        )

        df_relationships, df_unrelated, df_complex_shape_rooms = sort_and_update_dataframes(
            df_relationships, df_unrelated, df_complex_shape_rooms
        )

        return df_relationships, df_unrelated, df_complex_shape_rooms

    except Exception as e:
        logging.exception(f"Error in find_ceiling_room_relationships: {e}")
        raise

# def find_ceiling_room_relationships_old(room_elements, ceiling_elements, building_A_levels, building_B_levels, common_levels, prefix, timestamp):
#     """
#     Find the relationship between ceilings and rooms based on the two vectors.
    
#     Args:
#         room_elements (list): List of room elements.
#         ceiling_elements (list): List of ceiling elements.
#     Returns:
#         tuple: Two pandas DataFrames - one for related ceilings and rooms, one for unrelated ceilings.
#     """
#     try:
#         relationships = []
#         unrelated_ceilings = []
#         complex_shape_rooms = []
#         no_geometry_rooms = []
#         no_geometry_ceilings = []
        
#         building_levels = [building_A_levels, building_B_levels, common_levels]

#         start_time = datetime.datetime.now() # Start time for performance measurement

#         # notification_counter = 0

#         for i, ceiling in enumerate(ceiling_elements, 1):
#             ceiling_start_time = datetime.datetime.now() # Start time for performance measurement

#             ceiling_id = ceiling.Id
#             ceiling_details = get_ceiling_details(ceiling)
            
#             # Check if the ceiling has geometry
#             if not get_element_geometry(ceiling_id):
#                 logging.exception(f"No geometry found for ceiling ID: {ceiling_id.IntegerValue}")
#                 logging.exception(traceback.format_exc())
#                 unrelated_ceilings.append(ceiling_details)
#                 no_geometry_ceilings.append(ceiling_id)
#                 continue
            
#             direct_intersections = []
#             xy_intersections = []

#             # Iterate through all rooms to find intersections with the current ceiling
#             for j, room in enumerate(room_elements, 1):
#                 room_start_time = datetime.datetime.now() # Start time for performance measurement
#                 room_id = room.Id
#                 room_details = get_room_details(room)
                
#                 # Skip rooms without valid geometry
#                 if not get_element_geometry(room_id):
#                     if room_id not in no_geometry_rooms:
#                         no_geometry_rooms.append(room_id)
#                     continue
                
#                 # Check for direct intersection
#                 try:
#                     direct_intersection = check_direct_intersection(room_id, ceiling_id, prefix, timestamp)
#                     # if not direct_intersection:
#                     #     logging.exception(f"Direct intersection: {direct_intersection} for room ID: {room_id.IntegerValue}, ceiling ID: {ceiling_id.IntegerValue}")
#                     # print(f"Direct intersection: {direct_intersection}")
#                 except Exception as e:
#                     logging.exception(f"Error in direct intersection check: {e}")
#                     logging.exception(traceback.format_exc())
#                     direct_intersection = False
                
#                 # Check for XY projection intersection
#                 xy_intersection = project_and_check_xy_intersection(room_id, ceiling_id) if not direct_intersection else True
#                 # if not xy_intersection:
#                 #     logging.exception(f"XY projection intersection: {xy_intersection} for room ID: {room_id.IntegerValue}, ceiling ID: {ceiling_id.IntegerValue}")
                
#                 # If there's any kind of intersection, calculate the area and add to matched_rooms
#                 if direct_intersection or xy_intersection:
#                     intersection_area_3d, intersection_area_xy, is_complex_shape = calculate_intersection_areas(room_id, ceiling_id)
#                     # logging.exception(f"Intersection areas: 3D - {intersection_area_3d}, XY - {intersection_area_xy}, Complex shape - {is_complex_shape}")
#                     distance = delta_ceiling_above_room(room_id, ceiling_id)
#                     # logging.exception(f"Distance: {distance}")
#                     if is_complex_shape:
#                         complex_shape_rooms.append({
#                             'Room_ID': room_details[0],
#                             'Room_Name': room_details[1],
#                             'Room_Number': room_details[2],
#                             'Room_Level': room_details[3],
#                             'Room_Building': room_details[4],
#                             'Room_Area_sqm (* 1.05)': room_details[6] * 1.05,
#                             'Room_Ceiling_Finish': room_details[5],
#                             'Ceiling_ID': ceiling_details[0],
#                             'Ceiling_Level': ceiling_details[3],
#                             'Ceiling_Description': ceiling_details[1],
#                             'Ceiling_Description_Old': ceiling_details[1],
#                             'Ceiling_Area_sqm (* 1.3)': ceiling_details[2] * 1.3,
#                             'Intersection_Area_3D_sqm': intersection_area_3d,
#                             'Intersection_Area_XY_sqm': intersection_area_xy,
#                             'Intersection_Area_sqm': max(intersection_area_3d, intersection_area_xy),
#                             'Direct_Intersection': direct_intersection,
#                             'XY_Projection_Intersection': xy_intersection,
#                             'Distance': distance
#                         })
#                         continue  # Skip adding to relationships
#                     room_data = {
#                         'Ceiling_ID': ceiling_details[0],
#                         'Ceiling_Description': ceiling_details[1],
#                         'Ceiling_Description_Old': ceiling_details[1],
#                         'Room_Ceiling_Finish': room_details[5],
#                         'Ceiling_Area_sqm': ceiling_details[2],
#                         'Ceiling_Level': ceiling_details[3],
#                         'Room_ID': room_details[0],
#                         'Room_Name': room_details[1],
#                         'Room_Number': room_details[2],
#                         'Room_Level': room_details[3],
#                         'Room_Building': room_details[4],
#                         'Room_Area_sqm': room_details[6],
#                         'Intersection_Area_3D_sqm': intersection_area_3d,
#                         'Intersection_Area_XY_sqm': intersection_area_xy,
#                         'Intersection_Area_sqm': max(intersection_area_3d, intersection_area_xy),
#                         'Direct_Intersection': direct_intersection,
#                         'XY_Projection_Intersection': xy_intersection,
#                         'Distance': distance
#                     }
#                     if direct_intersection:
#                         direct_intersections.append(room_id)
#                         relationships.append(room_data) #immidiatly add direct intersections to the relationships
#                     elif xy_intersection:
#                         xy_intersections.append(room_data)

#                 # if notification_counter % 200 == 0:  
#                 #     elapsed_time = (datetime.datetime.now() - start_time).total_seconds()
#                 #     processing_time = (datetime.datetime.now() - room_start_time).total_seconds()
#                 #     ceiling_digits = len(str(len(ceiling_elements)))
#                 #     room_digits = len(str(len(room_elements)))
#                 #     print(f"Processing pair c{i:{ceiling_digits}d}/{len(ceiling_elements)}   r{j:{room_digits}d}/{len(room_elements)}")
                
#                 # notification_counter += 1
#                 elapsed_time = (datetime.datetime.now() - start_time).total_seconds()
#                 avg_processing_time = elapsed_time / i / j

#             # distances = [(room['Room_ID'], room['Distance']) for room in matched_rooms]

#             # If there are no direct intersections, process the XY intersections
#             if xy_intersections:
#                 # Filter out rooms with negative or zero distance
#                 positive_distance_rooms = [room for room in xy_intersections if room['Distance'] >= 0]
#                 # logging.exception(f"Positive distance rooms: {len(positive_distance_rooms)}")
#                 # Calculate level distances once
#                 level_distances = calculate_max_level_heights(doc, building_levels)

#                 if positive_distance_rooms:
#                     # Find the smallest positive distance
#                     min_positive_distance = min(room['Distance'] for room in positive_distance_rooms)
#                     room_with_min_distance = min(positive_distance_rooms, key=lambda x: x['Distance'])
#                     min_distance_level = doc.GetElement(ElementId(room_with_min_distance['Room_ID'])).LevelId
                    
#                     filtered_rooms = [
#                         room for room in positive_distance_rooms
#                         if doc.GetElement(ElementId(room['Room_ID'])).LevelId == min_distance_level
#                     ]

#                     leveled_rooms = []

#                     for room in filtered_rooms:
#                         if room['Room_Building'] == "A":
#                             # print(f"Room ID: {room['Room_ID']}, Number: {room['Room_Number']}, Name: {room['Room_Name']}, Distance: {room['Distance']}, Level: {room['Room_Level']}, Ceiling ID: {room['Ceiling_ID']}, Ceiling Level: {room['Ceiling_Level']}, max allowed: {level_distances[0]}")
#                             if room['Distance'] <= level_distances[0]:
#                                 leveled_rooms.append(room)
#                         if room['Room_Building'] == "B":
#                             # print(f"Room ID: {room['Room_ID']}, Number: {room['Room_Number']}, Name: {room['Room_Name']}, Distance: {room['Distance']}, Level: {room['Room_Level']}, Ceiling ID: {room['Ceiling_ID']}, Ceiling Level: {room['Ceiling_Level']}, max allowed: {level_distances[1]}")
#                             if room['Distance'] <= level_distances[1]:
#                                 leveled_rooms.append(room)
#                         else:
#                             # print(f"Room ID: {room['Room_ID']}, Number: {room['Room_Number']}, Name: {room['Room_Name']}, Distance: {room['Distance']}, Level: {room['Room_Level']}, Ceiling ID: {room['Ceiling_ID']}, Ceiling Level: {room['Ceiling_Level']}, max allowed: {level_distances[2]}")
#                             if room['Distance'] <= level_distances[2]:
#                                 leveled_rooms.append(room)
                    
#                     if not leveled_rooms:
#                         logging.exception(f"Ceiling ID {ceiling_id.IntegerValue} has no rooms in the same level with positive distance to it")
#                         continue
#                     else:
#                         relationships.extend(leveled_rooms)
#                 else:
#                     logging.exception(f"Ceiling ID {ceiling_id.IntegerValue} has no positive distance to any room in xy projections")

#             # If no rooms intersect with the ceiling, mark the ceiling as unrelated
#             if not direct_intersections and not xy_intersections:
#                 unrelated_ceilings.append(ceiling_details)

#             current_time = datetime.datetime.now()
#             elapsed_time = (current_time - start_time).total_seconds()
#             total_time_expected = elapsed_time / i * len(ceiling_elements)
#             percent = (elapsed_time / total_time_expected) * 100 if total_time_expected > 0 else 0
#             time_to_finish = total_time_expected - elapsed_time
#             print(f"Time to finish: {abs(time_to_finish):.2f}s ({percent:.2f}%);   Elapsed time: {elapsed_time:.2f}s;    Avg. processing time: {avg_processing_time:.4f}s")

#         # Collect debug messages for rooms without geometry
#         for room_id in no_geometry_rooms:
#             room = doc.GetElement(room_id)
#             room_name = room.get_Parameter(BuiltInParameter.ROOM_NAME).AsString()
#             room_number = room.get_Parameter(BuiltInParameter.ROOM_NUMBER).AsString()
#             room_building_param = room.LookupParameter("בניין")
#             room_building = room_building_param.AsString() if room_building_param else None
#             room_level = doc.GetElement(room.LevelId).Name
#             logging.exception(f"No geometry found for room ID: {room_id.IntegerValue} - Building: {room_building}, Level: {room_level}, Name: {room_name}, Number: {room_number}")

#         # Create DataFrames from the collected data
#         df_relationships = pd.DataFrame(relationships)
#         df_unrelated = pd.DataFrame(unrelated_ceilings, columns=['Ceiling_ID', 'Ceiling_Description', 'Ceiling_Area_sqm', 'Ceiling_Level'])
#         df_complex_shape_rooms = pd.DataFrame(complex_shape_rooms)

#         # Sort the relationships DataFrame
#         if not df_relationships.empty:
#             df_relationships = df_relationships.sort_values(
#                 by=['Room_Building', 'Room_Level', 'Room_Number', 'Ceiling_Description'],
#                 key=lambda x: x.map(custom_sort_key)
#             )
        
#         # Change Ceiling_Description to Room_Ceiling_Finish by rules, store the original value in Ceiling_Description_Old
#         if not df_relationships.empty:
#             df_relationships.loc[(df_relationships['Ceiling_Description'] == '"פלב"מ "ריגיד') & (df_relationships['Intersection_Area_sqm'] > 0) & (df_relationships['Intersection_Area_sqm'] / df_relationships['Ceiling_Area_sqm'] > 0.09) & (df_relationships['Room_Ceiling_Finish'] == 'פנל מבודד צבוע'), 'Ceiling_Description'] = 'פנל מבודד צבוע'
#             df_relationships.loc[(df_relationships['Ceiling_Description'] == "תקרת פח מחורר/אקופון 60/60 אדוונטיג' בשילוב סינורי גבס") & (df_relationships['Intersection_Area_sqm'] > 0) & (df_relationships['Intersection_Area_sqm'] / df_relationships['Ceiling_Area_sqm'] > 0.09) & (df_relationships['Room_Ceiling_Finish'] == 'פורניר  ע"ג MDF  גודל 60/60 בשילוב סינורי גבס'), 'Ceiling_Description'] = 'פורניר  ע"ג MDF  גודל 60/60 בשילוב סינורי גבס'
#             df_relationships.loc[(df_relationships['Ceiling_Description'] == "תקרת פח מחורר/אקופון 60/60 אדוונטיג' בשילוב סינורי גבס") & (df_relationships['Intersection_Area_sqm'] > 0) & (df_relationships['Intersection_Area_sqm'] / df_relationships['Ceiling_Area_sqm'] > 0.09) & (df_relationships['Room_Ceiling_Finish'] == 'תקרת פח מחורר 60/60 בשילוב סינורי גבס'), 'Ceiling_Description'] = 'תקרת פח מחורר 60/60 בשילוב סינורי גבס'

#         # Sort the unrelated ceilings DataFrame
#         if not df_unrelated.empty:
#             df_unrelated = df_unrelated.sort_values(
#                 by=['Ceiling_Level', 'Ceiling_Description'],
#                 key=lambda x: x.map(custom_sort_key)
#             )

#         # Sort the complex shape rooms DataFrame
#         if not df_complex_shape_rooms.empty:
#             df_complex_shape_rooms = df_complex_shape_rooms.sort_values(
#                 by=['Room_Building', 'Room_Level', 'Room_Number'],
#                 key=lambda x: x.map(custom_sort_key)
#             )

#         return df_relationships, df_unrelated, df_complex_shape_rooms
#     except Exception as e:
#         logging.exception(f"Error (internal) in find_ceiling_room_relationships: {e}")
#         logging.exception(traceback.format_exc())
#         raise

def find_rooms_without_ceilings(df_relationships, room_elements, df_complex_shape_rooms):
    """
    Identify rooms that don't have associated ceilings.
    
    Args:
        df_relationships (pandas.DataFrame): DataFrame containing ceiling-room relationships.
        room_elements (list): List of all room elements.
        df_complex_shape_rooms (pandas.DataFrame): DataFrame containing complex shape rooms.
    Returns:
        pandas.DataFrame: DataFrame containing rooms without ceilings.
    """
    if not df_relationships.empty:
        # Get all room IDs from the relationships DataFrame
        rooms_with_ceilings = set(df_relationships['Room_ID'])
        
        # Identify rooms without ceilings
        rooms_without_ceilings = []
        for room in room_elements:
            room_id = room.Id.IntegerValue
            if room_id not in rooms_with_ceilings:
                room_details = get_room_details(room)
                rooms_without_ceilings.append({
                    'Is_Complex_Shape': False,
                    'Room_Building': room_details[4],
                    'Room_Level': room_details[3],
                    'Room_Number': room_details[2],
                    'Room_Name': room_details[1],
                    'Room_Area_sqm': room_details[6],
                    'Room_ID': room_details[0]
                })
        
        rooms_without_ceilings_dict = {room['Room_ID']: room for room in rooms_without_ceilings}

        # Add a unique mention for each complex shape room
        for index, row in df_complex_shape_rooms.iterrows():
            room_id = row['Room_ID']
            if room_id not in rooms_without_ceilings_dict:
                rooms_without_ceilings_dict[room_id] = {
                    'Is_Complex_Shape': True,
                    'Room_Building': row['Room_Building'],
                    'Room_Level': row['Room_Level'],
                    'Room_Number': row['Room_Number'],
                    'Room_Name': row['Room_Name'],
                    'Room_Area_sqm': row['Room_Area_sqm'] if 'Room_Area_sqm' in row else (row['Room_Area_sqm (* 1.05)'] / 1.05 if 'Room_Area_sqm (* 1.05)' in row else None),
                    'Room_ID': row['Room_ID']
                }
            else:
                rooms_without_ceilings_dict[room_id]['Is_Complex_Shape'] = True
        
        rooms_without_ceilings = list(rooms_without_ceilings_dict.values())
        df_rooms_without_ceilings = pd.DataFrame(rooms_without_ceilings)
        
        # Sort the DataFrame
        df_rooms_without_ceilings = df_rooms_without_ceilings.sort_values(
            by=['Is_Complex_Shape', 'Room_Building', 'Room_Level', 'Room_Number'],
            key=lambda x: x.map(custom_sort_key)
        )

        return df_rooms_without_ceilings
    else:
        return pd.DataFrame()

def pivot_data(df_relationships):
    """
    Group the relationships data around rooms, preserving all ceiling information in separate rows.
    Handles DataFrames that may not contain all expected columns.
   
    Args:
        df_relationships (pandas.DataFrame): DataFrame containing ceiling-room relationships.
    Returns:
        pandas.DataFrame: Grouped DataFrame with rooms and all associated ceiling information.
    """
    # Define all expected columns
    expected_columns = ['Room_Building', 'Room_Level', 'Room_Number', 'Room_Name', 'Room_ID', 'Room_Area_sqm',
                        'Ceiling_ID', 'Ceiling_Description', 'Room_Ceiling_Finish', 'Ceiling_Area_sqm',
                        'Ceiling_Level', 'Intersection_Area_3D_sqm', 'Intersection_Area_XY_sqm',
                        'Intersection_Area_sqm', 'Direct_Intersection', 'XY_Projection_Intersection',
                        'Distance', 'Ceiling_Description_Old']
    
    # Add missing columns with NaN values
    for col in expected_columns:
        if col not in df_relationships.columns:
            df_relationships[col] = np.nan
    
    # Handle 'Room_Area_sqm (* 1.05)' column if present
    if 'Room_Area_sqm (* 1.05)' in df_relationships.columns:
        df_relationships['Room_Area_sqm'] = df_relationships['Room_Area_sqm (* 1.05)'] / 1.05
        df_relationships.drop('Room_Area_sqm (* 1.05)', axis=1, inplace=True)
    
    # Handle 'Ceiling_Area_sqm (* 1.3)' column if present
    if 'Ceiling_Area_sqm (* 1.3)' in df_relationships.columns:
        df_relationships['Ceiling_Area_sqm'] = df_relationships['Ceiling_Area_sqm (* 1.3)'] / 1.3
        df_relationships.drop('Ceiling_Area_sqm (* 1.3)', axis=1, inplace=True)
    
    # If DataFrame is empty, return it with all expected columns
    if df_relationships.empty:
        return pd.DataFrame(columns=expected_columns + ['Room', 'Ceilings_in_Room', 'Has_Gypsum_Ceiling'])
    
    # Sort the DataFrame
    df_sorted = df_relationships.sort_values(
        by=['Room_Building', 'Room_Level', 'Room_Number', 'Ceiling_ID'],
        key=lambda x: x.map(custom_sort_key)
    )
   
    # Add a column for the number of ceilings per room
    ceilings_per_room = df_sorted.groupby(['Room_ID'])['Ceiling_ID'].transform('count')
    df_sorted['Ceilings_in_Room'] = ceilings_per_room
   
    # Add helper new columns
    df_sorted['Room'] = df_sorted['Room_Number'].fillna('') + ' - ' + df_sorted['Room_Name'].fillna('')
    df_sorted['Has_Gypsum_Ceiling'] = df_sorted.groupby('Room_ID')['Ceiling_Description'].transform(
        lambda x: 1 if (x == 'תקרת גבס').any() else 0
    )
   
    # Adjust Intersection_Area_sqm
    df_sorted['Intersection_Area_sqm'] = df_sorted['Intersection_Area_sqm'].apply(lambda x: 0 if pd.notna(x) and x <= 0.36 else x)
    
    # Reorder columns for better readability
    columns_order = ['Room_Building', 'Room_Level', 'Room_Number', 'Room_Name', 'Room', 'Room_ID', 'Room_Area_sqm',
                    'Ceilings_in_Room', 'Has_Gypsum_Ceiling', 'Ceiling_ID', 'Ceiling_Description',
                    'Room_Ceiling_Finish', 'Ceiling_Area_sqm', 'Ceiling_Level', 'Intersection_Area_3D_sqm',
                    'Intersection_Area_XY_sqm', 'Intersection_Area_sqm', 'Direct_Intersection',
                    'XY_Projection_Intersection', 'Distance', 'Ceiling_Description_Old']
   
    df_grouped = df_sorted[columns_order]
   
    return df_grouped

def adjust_column_widths(ws):
    '''
    Adjust the column widths in the worksheet based on the content.

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet to adjust.
    '''
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

def apply_header_format(ws):
    '''
    Apply formatting to the header row in the worksheet.

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet to format.
    '''
    for cell in ws[1]:
        # Apply formatting to the header row
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
        cell.border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Freeze the header row
        ws.freeze_panes = 'A2'
        
        # Hide the ID columns
        for col in ws.columns:
            if "ID" in col[0].value:  # Check if "ID" is in the header cell
                col_letter = col[0].column_letter
                ws.column_dimensions[col_letter].hidden = True

def clear_all_lru_caches():
    for func in lru_cached_functions:
        func.cache_clear()

def main():
    clear_all_lru_caches()

    # Initialize timestamp
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

    # Collect file name for excel name prefix
    file_name = doc.PathName.split("/")[-1].split(".")[0].split("_")
    prefix = f"{file_name[0]}_{file_name[1]}"

    setup_logging(prefix, timestamp)

    try:
        output_dir = os.path.join("C:", "Mac", "Home", "Documents", "Shapir", "Exports", prefix, timestamp)
        os.makedirs(output_dir, exist_ok=True)
    except Exception as e:
        logging.exception(f"Error in creating output directory: {e}")

    try:
        # Collect room and ceiling elements
        room_elements = FilteredElementCollector(doc).OfClass(SpatialElement).OfCategory(BuiltInCategory.OST_Rooms).ToElements()
        ceiling_elements = FilteredElementCollector(doc).OfClass(Ceiling).ToElements()

        # For testing purposes, use a subset of rooms and ceilings
        # room_elements = [doc.GetElement(ElementId(1686235))]
        # ceiling_elements = [doc.GetElement(ElementId(2523734))]

        # Find the relationships between ceilings and rooms
        try:
            # C AB:
            building_A_levels = {ElementId(6533282), ElementId(694), ElementId(6568872)}  # 00, 01A, RF
            building_B_levels = {ElementId(8968108), ElementId(9048752), ElementId(9057595)}  # B 00, B 01, B RF
            common_levels = {ElementId(311)}  # -0.5

            # S AB:
            # building_A_levels = {}
            # building_B_levels = {}
            # common_levels = {ElementId(2003554), ElementId(13071), ElementId(15913), ElementId(1764693), ElementId(2102106)} # -1, 00, 01, 02, 03

            df_relationships, df_unrelated, df_complex_shape = find_ceiling_room_relationships(room_elements, ceiling_elements, building_A_levels, building_B_levels, common_levels, prefix, timestamp)
        except Exception as e:
            logging.exception(f"Error in find_ceiling_room_relationships: {e}")
            logging.exception(traceback.format_exc())

        # Pivot the relationships data
        try:
            df_grouped = pivot_data(df_relationships)
        except Exception as e:
            logging.exception(f"Error in pivot_data (relationships): {e}")
            logging.exception(traceback.format_exc())

        # Pivot the complex shape rooms data
        try:
            df_complex_shape = pivot_data(df_complex_shape)
        except Exception as e:
            logging.exception(f"Error in pivot_data (complex shape rooms): {e}")
            logging.exception(traceback.format_exc())

        # Find rooms without ceilings
        try:
            df_rooms_without_ceilings = find_rooms_without_ceilings(df_relationships, room_elements, df_complex_shape)
        except Exception as e:
            logging.exception(f"Error in find_rooms_without_ceilings: {e}")
            logging.exception(traceback.format_exc())
        
        # Output the dataframe to Excel with timestamp
        output_file_path = f"C:\\Mac\\Home\\Documents\\Shapir\\Exports\\{prefix}\\{timestamp}\\ceiling_room_relationships.xlsx"

        try:
            # Export to Excel with formatting
            with pd.ExcelWriter(output_file_path, engine='xlsxwriter') as writer:
                df_grouped.to_excel(writer, sheet_name='Ceiling-Room Relationships', index=False)
                df_unrelated.to_excel(writer, sheet_name='Unrelated Ceilings', index=False)
                df_rooms_without_ceilings.to_excel(writer, sheet_name='Rooms Without Ceilings', index=False)
                df_complex_shape.to_excel(writer, sheet_name='Complex Shape Rooms', index=False)


                workbook = writer.book
                no_ceiling_worksheet = writer.sheets['Rooms Without Ceilings']
                unrelated_worksheet = writer.sheets['Unrelated Ceilings']
                grouped_worksheet = writer.sheets['Ceiling-Room Relationships']
                complex_shape_worksheet = writer.sheets['Complex Shape Rooms']
                
            # Load the saved workbook and apply formatting
            wb = load_workbook(output_file_path)
            
            # Apply header format
            for sheet_name in wb.sheetnames:
                apply_header_format(wb[sheet_name])

            # Adjust column widths
            for sheet_name in wb.sheetnames:
                adjust_column_widths(wb[sheet_name])
        except Exception as e:
            logging.exception(f"Error in saving to Excel: {e}")
            logging.exception(traceback.format_exc())

        # Save the workbook
        try:
            wb.save(output_file_path)
            print(f"Schedule saved to {output_file_path}")
        except Exception as e:
            logging.exception(f"Error in saving workbook: {e}")
            logging.exception(traceback.format_exc())

    except Exception as e:
        logging.exception(f"Error in main function: {e}")
        logging.exception(traceback.format_exc())
    
    clear_all_lru_caches()

# Call the main function
if __name__ == "__main__":
    main()
