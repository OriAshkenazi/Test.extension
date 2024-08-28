#! python3

# import clr
# clr.AddReference('RevitAPI')
# from Autodesk.Revit.DB import *

# doc = __revit__.ActiveUIDocument.Document

# def print_level_names_and_ids(doc):
#     # Collect all levels in the document
#     levels = FilteredElementCollector(doc).OfClass(Level).ToElements()
    
#     # Sort levels by elevation
#     levels = sorted(levels, key=lambda l: l.Elevation)
    
#     # Print tuples of level names and IDs
#     print("Level Name and ID tuples:")
#     for level in levels:
#         name = level.Name
#         id = level.Id.IntegerValue
#         print(f"('{name}', {id})")


# print_level_names_and_ids(__revit__.ActiveUIDocument.Document)

# print(doc.PathName)

# import clr
# clr.AddReference('RevitAPI')
# from Autodesk.Revit.DB import *

# def safe_get_property(obj, prop_name):
#     try:
#         return getattr(obj, prop_name)
#     except Exception as e:
#         return f"Error accessing {prop_name}: {str(e)}"

# def debug_compound_ceiling(doc, element_id):
#     """
#     Debug function to investigate issues with compound ceiling elements.
    
#     Args:
#         doc (Document): The Revit document.
#         element_id (int): The ID of the compound ceiling element to debug.
#     """
#     try:
#         # Get the element
#         element = doc.GetElement(ElementId(element_id))
#         if not element:
#             print(f"No element found with ID {element_id}")
#             return

#         print(f"Debugging Compound Ceiling Element (ID: {element_id})")
#         print("-" * 50)

#         # Basic element info
#         print(f"Element Type: {type(element)}")
#         print(f"Category: {safe_get_property(element, 'Category').Name if safe_get_property(element, 'Category') else 'No Category'}")

#         # Check if it's a ceiling
#         print(f"Is Ceiling instance: {isinstance(element, Ceiling)}")

#         # Get element type
#         element_type = doc.GetElement(element.GetTypeId())
#         print(f"Element Type: {type(element_type)}")
#         print(f"Element Type ID: {element.GetTypeId()}")
#         print(f"Element Type Name: {safe_get_property(element_type, 'Name')}")

#         # Try to get family information
#         family = safe_get_property(element_type, 'Family')
#         print(f"Family: {family}")
#         if family:
#             print(f"Family Name: {safe_get_property(family, 'Name')}")

#         # Check for compound structure
#         if hasattr(element_type, 'GetCompoundStructure'):
#             try:
#                 compound_structure = element_type.GetCompoundStructure()
#                 if compound_structure:
#                     print("Compound Structure found:")
#                     layers = compound_structure.GetLayers()
#                     for i, layer in enumerate(layers):
#                         material = doc.GetElement(layer.MaterialId)
#                         material_name = safe_get_property(material, 'Name') if material else "No Material"
#                         print(f"  Layer {i+1}: Thickness = {layer.Width}, Material = {material_name}")
#                 else:
#                     print("No Compound Structure found")
#             except Exception as e:
#                 print(f"Error accessing Compound Structure: {str(e)}")
#         else:
#             print("GetCompoundStructure method not found")

#         # Try to get area
#         area_param = element.get_Parameter(BuiltInParameter.HOST_AREA_COMPUTED)
#         if area_param:
#             area = area_param.AsDouble() * 0.092903  # Convert to square meters
#             print(f"Area: {area:.2f} m²")
#         else:
#             print("No HOST_AREA_COMPUTED parameter found")
        
#         # Try to get volume
#         volume_param = element.get_Parameter(BuiltInParameter.HOST_VOLUME_COMPUTED)
#         if volume_param:
#             volume = volume_param.AsDouble() * 0.0283168  # Convert to cubic meters
#             print(f"Volume: {volume:.2f} m³")
#         else:
#             print("No HOST_VOLUME_COMPUTED parameter found")

#         # List all parameters
#         print("\nAll Parameters:")
#         for param in element.Parameters:
#             try:
#                 if param.HasValue:
#                     if param.StorageType == StorageType.Double:
#                         value = param.AsDouble()
#                     elif param.StorageType == StorageType.Integer:
#                         value = param.AsInteger()
#                     elif param.StorageType == StorageType.String:
#                         value = param.AsString()
#                     elif param.StorageType == StorageType.ElementId:
#                         value = param.AsElementId()
#                     else:
#                         value = "Unknown"
#                     print(f"  {param.Definition.Name}: {value}")
#             except Exception as e:
#                 print(f"  Error accessing parameter {param.Definition.Name}: {str(e)}")

#     except Exception as e:
#         print(f"Error in debug_compound_ceiling: {str(e)}")
#         import traceback
#         print(traceback.format_exc())

# # Usage
# doc = __revit__.ActiveUIDocument.Document  # Get the active document
# debug_compound_ceiling(doc, 10476805)  # Call the debug function with the specific element ID

# import clr
# clr.AddReference('RevitAPI')
# clr.AddReference('RevitAPIUI')
# from Autodesk.Revit.DB import *
# from Autodesk.Revit.UI import *

# doc = __revit__.ActiveUIDocument.Document

# def get_view_template_info(template_name):
#     # Find the view template
#     collector = FilteredElementCollector(doc).OfClass(View)
#     template = next((v for v in collector if v.IsTemplate and v.Name == template_name), None)
    
#     if not template:
#         print(f"View template '{template_name}' not found.")
#         return
    
#     info = []
#     info.append(f"View Template: {template_name}\n")
    
#     # Basic properties
#     info.append(f"Detail Level: {template.DetailLevel}")
#     info.append(f"Visual Style: {template.DisplayStyle}")
    
#     # Visibility/Graphics overrides
#     visible_categories = []
#     hidden_categories = []
#     transparency_settings = []
    
#     for category in doc.Settings.Categories:
#         if category.CategoryType == CategoryType.Model:
#             is_visible = not template.GetCategoryHidden(category.Id)
#             if is_visible:
#                 visible_categories.append(category.Name)
#             else:
#                 hidden_categories.append(category.Name)
            
#             overrides = template.GetCategoryOverrides(category.Id)
#             if overrides.Transparency > 0:
#                 transparency_settings.append(f"{category.Name}: {overrides.Transparency}%")
    
#     info.append("\nVisible Categories:")
#     info.extend(f"- {cat}" for cat in sorted(visible_categories))
    
#     info.append("\nHidden Categories:")
#     info.extend(f"- {cat}" for cat in sorted(hidden_categories))
    
#     if transparency_settings:
#         info.append("\nTransparency Settings:")
#         info.extend(transparency_settings)
    
#     # Filters
#     filters = template.GetFilters()
#     if filters:
#         info.append("\nFilters:")
#         for filter_id in filters:
#             filter_element = doc.GetElement(filter_id)
#             info.append(f"- {filter_element.Name}")
    
#     # Other properties
#     info.append("\nOther Properties:")
#     for param in template.Parameters:
#         if param.HasValue:
#             if param.StorageType == StorageType.String:
#                 value = param.AsString()
#             elif param.StorageType == StorageType.Double:
#                 value = param.AsDouble()
#             elif param.StorageType == StorageType.Integer:
#                 value = param.AsInteger()
#             else:
#                 value = "Unable to retrieve"
#             info.append(f"- {param.Definition.Name}: {value}")
    
#     return "\n".join(info)

# # Usage
# template_name = "VDC - HVAC ELEC verification"
# template_info = get_view_template_info(template_name)

# if template_info:
#     print(template_info)
    
#     # Optionally, save to a file
#     with open("view_template_info.txt", "w") as f:
#         f.write(template_info)
#     print("\nInformation has been saved to 'view_template_info.txt'")
# else:
#     print("Failed to retrieve template information.")

import clr
import sys
import os
from collections import defaultdict

clr.AddReference('RevitAPI')
clr.AddReference('RevitAPIUI')
from Autodesk.Revit.DB import *
from Autodesk.Revit.UI import *

import openpyxl
from openpyxl.styles import Font
from tqdm import tqdm

doc = __revit__.ActiveUIDocument.Document
uidoc = __revit__.ActiveUIDocument

def get_linked_documents():
    linked_docs = []
    collector = FilteredElementCollector(doc).OfClass(RevitLinkInstance)
    for link_instance in collector:
        linked_doc = link_instance.GetLinkDocument()
        if linked_doc:
            linked_docs.append((linked_doc, link_instance.GetTotalTransform()))
    return linked_docs

def get_mech_equipment(document):
    return FilteredElementCollector(document).OfCategory(BuiltInCategory.OST_MechanicalEquipment).WhereElementIsNotElementType().ToElements()

def get_element_info(element, document, transform=None):
    location = element.Location.Point if element.Location and hasattr(element.Location, 'Point') else None
    if location and transform:
        location = transform.OfPoint(location)
    
    info = {
        "Document": document.Title,
        "WorksetId": element.WorksetId.IntegerValue,
        "UniqueId": element.UniqueId,
        "Id": element.Id.IntegerValue,
        "Name": element.Name,
        "Category": element.Category.Name,
        "FamilyName": element.Symbol.FamilyName if hasattr(element, 'Symbol') else "N/A",
        "TypeName": element.Name,
        "TypeId": element.GetTypeId().IntegerValue,
        "Location": f"({location.X:.2f}, {location.Y:.2f}, {location.Z:.2f})" if location else "N/A",
        "Mark": element.get_Parameter(BuiltInParameter.ALL_MODEL_MARK).AsString() if element.get_Parameter(BuiltInParameter.ALL_MODEL_MARK) else "N/A",
        "Comments": element.get_Parameter(BuiltInParameter.ALL_MODEL_INSTANCE_COMMENTS).AsString() if element.get_Parameter(BuiltInParameter.ALL_MODEL_INSTANCE_COMMENTS) else "N/A"
    }
    
    # Add all parameters
    for param in element.Parameters:
        if param.HasValue:
            if param.StorageType == StorageType.String:
                info[f"Param_{param.Definition.Name}"] = param.AsString()
            elif param.StorageType == StorageType.Double:
                info[f"Param_{param.Definition.Name}"] = param.AsDouble()
            elif param.StorageType == StorageType.Integer:
                info[f"Param_{param.Definition.Name}"] = param.AsInteger()
            elif param.StorageType == StorageType.ElementId:
                info[f"Param_{param.Definition.Name}"] = param.AsElementId().IntegerValue

    return info

def write_to_excel(data, sheet_name, workbook):
    sheet = workbook.create_sheet(sheet_name)
    
    # Write headers
    headers = list(data[0].keys()) if data else []
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    # Write data
    for row, item in enumerate(data, start=2):
        for col, (key, value) in enumerate(item.items(), start=1):
            sheet.cell(row=row, column=col, value=str(value))

def find_matches(local_elements, linked_elements):
    matches = []
    for local_elem in tqdm(local_elements, desc="Finding matches"):
        for linked_elem in linked_elements:
            if (local_elem["Name"] == linked_elem["Name"] and 
                local_elem["TypeId"] == linked_elem["TypeId"]):
                matches.append({
                    "Local_Document": local_elem["Document"],
                    "Local_WorksetId": local_elem["WorksetId"],
                    "Local_UniqueId": local_elem["UniqueId"],
                    "Local_Id": local_elem["Id"],
                    "Local_Name": local_elem["Name"],
                    "Local_TypeId": local_elem["TypeId"],
                    "Local_Location": local_elem["Location"],
                    "Linked_Document": linked_elem["Document"],
                    "Linked_WorksetId": linked_elem["WorksetId"],
                    "Linked_UniqueId": linked_elem["UniqueId"],
                    "Linked_Id": linked_elem["Id"],
                    "Linked_Name": linked_elem["Name"],
                    "Linked_TypeId": linked_elem["TypeId"],
                    "Linked_Location": linked_elem["Location"]
                })
    return matches

def main():
    output_file = r"C:\Temp\MechanicalEquipmentAnalysis.xlsx"
    workbook = openpyxl.Workbook()

    # Get project base point
    base_point = FilteredElementCollector(doc).OfCategory(BuiltInCategory.OST_ProjectBasePoint).FirstElement()
    if base_point:
        origin = XYZ(
            base_point.get_Parameter(BuiltInParameter.BASEPOINT_EASTWEST_PARAM).AsDouble(),
            base_point.get_Parameter(BuiltInParameter.BASEPOINT_NORTHSOUTH_PARAM).AsDouble(),
            base_point.get_Parameter(BuiltInParameter.BASEPOINT_ELEVATION_PARAM).AsDouble()
        )
    else:
        origin = XYZ(0, 0, 0)

    # 1. Get all mechanical equipment elements from the local model
    local_mech_equipment = list(tqdm(get_mech_equipment(doc), desc="Processing local elements"))
    local_info = [get_element_info(elem, doc) for elem in local_mech_equipment]
    print(f"Found {len(local_mech_equipment)} mechanical equipment elements in the local model.")
    
    # 2. Get all mechanical equipment elements from linked models
    linked_docs = get_linked_documents()
    linked_mech_equipment = []
    for linked_doc, transform in tqdm(linked_docs, desc="Processing linked documents"):
        linked_elements = get_mech_equipment(linked_doc)
        linked_mech_equipment.extend([(elem, linked_doc, transform) for elem in linked_elements])
    linked_info = [get_element_info(elem, doc, transform) for elem, doc, transform in linked_mech_equipment]
    print(f"Found {len(linked_mech_equipment)} mechanical equipment elements in linked models.")

    # Adjust locations
    for info in local_info + linked_info:
        if info["Location"] != "N/A":
            x, y, z = map(float, info["Location"][1:-1].split(", "))
            adjusted_location = XYZ(x - origin.X, y - origin.Y, z - origin.Z)
            info["Location"] = f"({adjusted_location.X:.2f}, {adjusted_location.Y:.2f}, {adjusted_location.Z:.2f})"

    # Write data to Excel
    write_to_excel(local_info, "Local Elements", workbook)
    write_to_excel(linked_info, "Linked Elements", workbook)

    # Find and write matches
    matches = find_matches(local_info, linked_info)
    write_to_excel(matches, "Potential Matches", workbook)

    # Remove default sheet
    workbook.remove(workbook['Sheet'])

    # Save the workbook
    workbook.save(output_file)
    print(f"Data saved to {output_file}")

if __name__ == '__main__':
    main()