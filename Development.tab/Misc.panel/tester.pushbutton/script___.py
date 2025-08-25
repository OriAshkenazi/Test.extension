#! python3
# -*- coding: utf-8 -*-
__title__ = "Export Section Dimensions"
__author__ = "Ori Ashkenazi"

import os
import pickle
from datetime import datetime
from pyrevit import revit, script
from Autodesk.Revit.DB import (
    FilteredElementCollector,
    Dimension,
    UnitUtils,
    UnitTypeId,
    ViewType
)
from openpyxl import load_workbook, Workbook
from openpyxl.comments import Comment

# Define path storage
settings_dir = os.path.join(os.path.expanduser("~"), "Documents", "SectionDimensions")
os.makedirs(settings_dir, exist_ok=True)
settings_path = os.path.join(settings_dir, "export_path.pkl")

def get_default_excel_path():
    filename = "SectionDimensions.xlsx"
    return os.path.join(settings_dir, filename)

# Load or define path
if os.path.exists(settings_path):
    with open(settings_path, "rb") as f:
        excel_path = pickle.load(f)
else:
    excel_path = get_default_excel_path()
    with open(settings_path, "wb") as f:
        pickle.dump(excel_path, f)

# Revit setup
doc = revit.doc
view = doc.ActiveView

if view.ViewType != ViewType.Section:
    script.get_output().print_md("❌ Please open a section view to run this script.")
    script.exit()

view_name = view.Name.strip().replace(":", "_").replace("/", "_")
collector = FilteredElementCollector(doc, view.Id).OfClass(Dimension)

# Classify and collect dimension values
def is_horizontal(v):
    # Accept either X- or Y-aligned as horizontal (perpendicular to vertical Z)
    return abs(v.Z) < 0.1 and (abs(v.X) > 0.5 or abs(v.Y) > 0.5)

def is_vertical(v):
    return abs(v.Z) > 0.5 and abs(v.X) < 0.1 and abs(v.Y) < 0.1

horizontal, vertical = [], []

output = script.get_output()
output.print_md("🔍 **Processing dimensions in view `{}`**".format(view_name))

for dim in collector:
    try:
        curve = dim.Curve
        dir = curve.Direction
        origin = curve.Origin
        val = dim.Value
        x = origin.X

        dir_str = f"Direction: X={dir.X:.2f}, Y={dir.Y:.2f}, Z={dir.Z:.2f}, Value={val:.2f}"

        if is_horizontal(dir):
            horizontal.append((x, val))
            output.print_md(f"📏 Horizontal ✓ → {dir_str}")
        elif is_vertical(dir):
            vertical.append((x, val))
            output.print_md(f"📐 Vertical ✓ → {dir_str}")
        else:
            output.print_md(f"❓ Skipped (neither horiz/vert) → {dir_str}")
    except Exception as e:
        output.print_md(f"⚠️ Error processing dimension: {e}")

horizontal.sort(key=lambda x: x[0])
vertical.sort(key=lambda x: x[0])

output.print_md(f"📊 Found {len(horizontal)} horizontal and {len(vertical)} vertical dimensions.")

def to_cm(feet):
    return round(UnitUtils.ConvertFromInternalUnits(feet, UnitTypeId.Centimeters), 1)

horizontal_cm = [to_cm(v) for _, v in horizontal]
vertical_cm = [to_cm(v) for _, v in vertical]

# Excel handling
if os.path.exists(excel_path):
    wb = load_workbook(excel_path)
else:
    wb = Workbook()
    # Remove default sheet
    default_sheet = wb.active
    if default_sheet.title == "Sheet":
        wb.remove(default_sheet)

# Remove existing sheet
if view_name in wb.sheetnames:
    std = wb[view_name]
    wb.remove(std)

ws = wb.create_sheet(view_name)

# Add timestamp comment
timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
ws["A1"].comment = Comment("Exported: {}".format(timestamp), "pyRevit")

# Headers
ws["B1"] = "Horizontal (cm)"
ws["C1"] = "Vertical (cm)"

# Populate data
for i in range(max(len(horizontal_cm), len(vertical_cm))):
    if i < len(horizontal_cm):
        ws.cell(row=i+2, column=2, value=horizontal_cm[i])
    if i < len(vertical_cm):
        ws.cell(row=i+2, column=3, value=vertical_cm[i])

wb.save(excel_path)

# Output
output = script.get_output()
output.print_md("✅ **Dimensions exported to Excel**")
output.print_md("📄 File: `{}`".format(excel_path))
output.print_md("📑 Sheet: `{}`".format(view_name))
output.print_md("📆 Timestamp: `{}`".format(timestamp))
output.print_md("✏️ To change output file, delete or edit:")
output.print_md("`{}`".format(settings_path))
